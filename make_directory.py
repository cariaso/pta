#!/usr/bin/env python

import codecs
import hashlib
import os
import pathlib
import re
import smtplib
import sys
import tempfile
import traceback
from copy import copy
from email import encoders
from email.mime.base import MIMEBase
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText

import click
import qrcode
import qrcode.image.svg
import reportlab.platypus
from reportlab.lib import colors
from reportlab.lib.enums import TA_RIGHT

# from reportlab.lib.pagesizes import A6, letter
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import inch
from reportlab.platypus import Image, PageBreak, Paragraph, Spacer, Table
from reportlab.platypus.doctemplate import BaseDocTemplate, PageTemplate
from reportlab.platypus.flowables import BalancedColumns, KeepTogether
from reportlab.platypus.frames import Frame
from reportlab.platypus.tableofcontents import TableOfContents


class MyDocTemplate(BaseDocTemplate):
    def __init__(self, filename, **kw):
        self.allowSplitting = 0
        BaseDocTemplate.__init__(
            self,
            filename,
            pagesize=(5.5 * inch, 8.5 * inch),
            **kw,
        )
        template = PageTemplate(
            "normal",
            [
                Frame(
                    x1=0.25 * inch,
                    y1=0.25 * inch,
                    width=5 * inch,
                    height=8 * inch,
                    id="Frame1",
                    # leftPadding=0.1 * inch,
                    # showBoundary=1,
                    showBoundary=0,
                )
            ],
            onPage=AllPageSetup,
        )
        self.addPageTemplates(template)

    def afterFlowable(self, flowable):
        "Registers TOC entries."
        if flowable.__class__.__name__ == "Paragraph":
            text = flowable.getPlainText()
            style = flowable.style.name

            if style == "TOCHeading1":
                level = 0
            elif style == "TOCHeading2":
                level = 1
            else:
                return

            E = [level, text, self.page]
            # if we have a bookmark name append that to our notify data
            bn = getattr(flowable, "_bookmarkName", None)
            if bn is not None:
                E.append(bn)
            self.notify("TOCEntry", tuple(E))


def AllPageSetup(canvas, doc):

    canvas.saveState()

    canvas.setAuthor("Somerset ES PTA")
    canvas.setTitle("Somerset ES 2024-2025 Directory")

    if doc.page == 1:
        image_path = "somerset_es_directory_cover.jpg"
        image_path = "somerset_es_directory_cover2.jpg"
        image_path = "somerset-staff-photo-23-24-cropped.jpg"
        image_path = "somerset_es_mary_vinograd.jpg"
        page_width, page_height = canvas._pagesize
        canvas.drawImage(
            image_path,
            0,
            1.75 * inch,
            width=5.5 * inch,
            height=4 * inch,
            preserveAspectRatio=True,
        )

        c = canvas

        # Cover Page Text with Drop Shadow
        shadow_offset = 0.025 * inch
        c.setFillColorRGB(0, 0, 0)
        c.setFont("Helvetica", 60)
        c.drawString(
            shadow_offset + 0.3 * inch, -shadow_offset + 7.25 * inch, "Somerset ES"
        )
        c.drawString(
            shadow_offset + 1 * inch, -shadow_offset + 6.25 * inch, "Directory"
        )

        pos3_y = 0.75 * inch
        c.drawString(shadow_offset + 0.8 * inch, -shadow_offset + pos3_y, "2024-2025")

        c.setFillColorRGB(102 / 256, 153 / 256, 102 / 256)
        c.drawString(0.3 * inch, 7.25 * inch, "Somerset ES")
        c.drawString(1 * inch, 6.25 * inch, "Directory")
        c.drawString(0.8 * inch, pos3_y, "2024-2025")

        ## Draw a line
        # c.setStrokeColorRGB(0,1,0.3) #choose your line color
        # c.line(2,2,2*inch,2*inch)

        ## Draw a rectangle
        # c.setFillColorRGB(1,1,0) #choose fill colour
        # c.rect(4*inch,4*inch,2*inch,3*inch, fill=1) #draw rectangle

    else:

        if doc.page > 4:
            canvas.drawCentredString(2.75 * inch, 0.2 * inch, "Page %d" % (doc.page))
        if hasattr(doc, "owner"):
            canvas.setSubject(doc.owner)
            # canvas.drawString(0.5 * inch, 0.5 * inch, doc.owner)

            canvas.rotate(90)
            fs = canvas._fontsize
            canvas.translate(1, -fs / 1.2)  # canvas._leading?
            canvas.drawString((3 + (doc.page / 100)) * inch, -0.05 * inch, doc.owner)

        # header
        # canvas.drawString(0.5 * inch, 8 * inch, doc.fund)
        # canvas.drawRightString(10.5 * inch, 8 * inch, doc.report_info)

        # footers

    canvas.restoreState()


@click.group()
def cli():
    pass


@cli.command("make-all-pdfs")
@click.option("--src", help="MCPS export .xlsx", required=True)
@click.option(
    "--pages/-no-pages",
    default=False,
    help="prepare N 1 page pdfs, in addition to the single N page pdf",
)
@click.option("--send/-no-send", default=False, help="send emails")
@click.option(
    "--board/-no-board", default=False, help="prepare versions for PTA board members"
)
@click.option(
    "--staff/-no-staff", default=False, help="prepare versions for staff members"
)
@click.option(
    "--parents/-no-parents", default=False, help="prepare versions for parents"
)
@click.pass_context
def make_all_pdfs(
    ctx, src, board=False, staff=False, parents=False, pages=None, send=False
):
    """setup whatever is needed"""

    pool = xlsx_to_pool(src)
    story = pool_to_story(pool)
    single_pdf = "somerset_directory.pdf"

    if pages:
        story_to_pdf(story, filename=single_pdf)

        from PyPDF2 import PdfReader, PdfWriter

        inputpdf = PdfReader(open(single_pdf, "rb"))

        for i in range(len(inputpdf.pages)):
            output = PdfWriter()
            output.add_page(inputpdf.pages[i])
            with open(
                f"pages/somerset-es-directory-page{i:05d}.pdf", "wb"
            ) as outputStream:
                output.write(outputStream)
    else:

        story_to_pdf(story)

    # do_filter = False

    pta_board = [
        "victoria.levitas@gmail.com",
        "Rachel Marx Boufford <rachel.boufford@gmail.com>",
        "Kate Julian <katejulian@yahoo.com>",
        "Deborah Leitner <debbie.leitner@outlook.com>",
        "Michael Cariaso <cariaso@gmail.com>",
        "Stacey Band <staceydwolf@gmail.com>",
        "Sharee Calverley Lawler <sharee.lawler@gmail.com>",
        "Steve Katrivanos <katrivanos@me.com>",
        "rebeca lamadrid <rebeca.lamadrid@gmail.com>",
        "DIANA VINUEZA <dianaximenav@gmail.com",
        "Chris Press <chris.press@gmail.com",
        # "Travis_J_Wiebe@mcpsmd.org",
        # "Bess_W_Treat@mcpsmd.org",
    ]

    pta_board = [
        "cariaso@gmail.com",
    ]

    login_username = "directory@somersetpta.org"
    from dotenv import load_dotenv

    load_dotenv()
    password = os.environ["SOMERSETPTA_DIRECTORY_PASSWORD"]

    if board:
        for owner in pta_board:
            story = pool_to_story(pool)
            safe_owner = make_filename_safe(owner)
            filename = f"unfiltered/somerset_directory_for_{safe_owner}.pdf"
            print(owner, filename)
            story_to_pdf(
                story,
                owner=owner,
                filename=filename,
            )

            stream = as_email(
                username=login_username,
                recipients=[owner],
                attachment=filename,
            )
            if send:
                send_emails(username=login_username, password=password, messages=stream)

    if staff:
        for staff_member in staff_order:
            owner = staff_member.get("email")
            if owner:
                story = pool_to_story(pool)
                safe_owner = make_filename_safe(owner)
                filename = f"unfiltered/somerset_directory_for_{safe_owner}.pdf"
                print(owner, filename)
                story_to_pdf(
                    story,
                    owner=owner,
                    filename=filename,
                )

    if parents:
        emails = xlsx_to_emails(src)
        for owner, students in emails.items():
            if "levitas" not in owner.lower():
                continue
            print(owner)

            if owner:
                # if do_filter:
                #     filtered_pool = filter_pool_to_students(pool, students)
                # else:
                #     filtered_pool = pool

                # story = pool_to_story(filtered_pool)
                # safe_owner = make_filename_safe(owner)
                # filename = f"filtered/filtered_somerset_directory_for_{safe_owner}.pdf"
                # print(owner, filename)
                # story_to_pdf(
                #     story,
                #     owner=owner,
                #     filename=filename,
                # )

                story = pool_to_story(pool)
                safe_owner = make_filename_safe(owner)
                filename = f"unfiltered/somerset_directory_for_{safe_owner}.pdf"
                print(owner, filename)
                story_to_pdf(
                    story,
                    owner=owner,
                    filename=filename,
                )

                stream = as_email(
                    username=login_username,
                    recipients=[owner],
                    # subject=subject,
                    # body=body,
                    attachment=filename,
                )
                if send:
                    send_emails(
                        username=login_username, password=password, messages=stream
                    )


@cli.command("refresh")
@click.option("--src", help="MCPS export .xlsx", required=False)
@click.option(
    "--parents/-no-parents", default=False, help="prepare versions for parents"
)
@click.pass_context
def refresh_the_pdf(ctx, src, board=False, staff=False, parents=False, pages=None):
    """live from google sheet to google drive"""

    src_url = "https://docs.google.com/spreadsheets/d/1YdQkan1JDiUqyGh30ugO-TuVnr1xySHcCi1n9Tlx-oc/edit"
    import gspread

    # contents of
    # /home/cariaso/.config/gspread/credentials.json
    # are real

    creds_path = "/home/cariaso/.config/gspread/credentials.json"
    gc = gspread.oauth(
        credentials_filename=creds_path,
    )

    sh = gc.open_by_url(src_url)
    worksheet = sh.get_worksheet(0)

    pool = xlsx_to_pool(src, sheet=worksheet)
    story = pool_to_story(pool)
    single_pdf = "somerset_directory.pdf"

    story_to_pdf(story, filename=single_pdf)


def make_filename_safe(filename):
    if "@" in filename:
        filename = filename.replace("@", "_at_")

    out = "".join(
        c for c in filename if c.isalpha() or c.isdigit() or c in " _"
    ).rstrip()
    return out


def get_teacher(record):
    teacher = record.get("Homeroom Teacher")
    if teacher is None:
        teacher = "No Teacher Set"
        print(" Warning no 'Homeroom Teacher' set for", record)
    return teacher


def get_grade(record):
    grade = record.get("Grade")
    if grade is None:
        grade = "No Grade Set"
        print(" Warning no Grade set for", record)
    return grade


def pool_to_teacher_grade(pool):
    out = {}
    for entry in pool:
        grade = get_grade(entry)
        teacher = get_teacher(entry)
        if grade not in out:
            out[grade] = {}
        if teacher not in out[grade]:
            out[grade][teacher] = []
        out[grade][teacher].append(entry)

    sorted_out = {
        "SE PreK": {},
        "K": {},
    }
    for grade in sorted(out):
        sorted_out[grade] = {}
        if None in out[grade]:
            breakpoint()
        for teacher in sorted(out[grade]):
            sorted_out[grade][teacher] = out[grade][teacher]
    return sorted_out


def pool_to_teacher_grade_student_uids(pool):
    out = {}
    tg = pool_to_teacher_grade(pool)
    for grade in tg:
        out[grade] = {}
        for teacher in tg[grade]:
            student_uids = []
            for entry in tg[grade][teacher]:
                astudent_uid = student_uid(entry)
                if astudent_uid not in student_uids:
                    student_uids.append(astudent_uid)
            out[grade][teacher] = student_uids
    return out


def get_street(address1):
    if address1 is None:
        out = None
    else:
        out = "unknown"
    src = address1

    if src:
        src = re.sub(r" Unit.*$", "", src)
        src = re.sub(r" Apt.*$", "", src)
        src = re.sub(r" Floor.*$", "", src)
        src = re.sub(r" Ste.*$", "", src)
        src = re.sub(r" Suite.*$", "", src)
        src = re.sub(r" #.*$", "", src)
        src = re.sub(r" Unit.*$", "", src)
    if src:
        m = re.search(r"^\d{1,} ([a-zA-Z0-9\s]+)", src)
        if m:
            out = m.group(1)
        else:
            pass

    # print(f"address1: [{address1}] street: [{out}]")
    return out


def student_uid(entry):
    student_ustr = entry.get("Student ID")
    if not student_ustr:
        student_name = entry.get("Student")
        dob = entry.get("Birth Date")
        # grade = entry.get("Grade")
        # teacher = entry.get("Homeroom Teacher")
        student_ustr = student_name + str(dob)

    uid = hashlib.sha1(student_ustr.encode("utf-8")).hexdigest()
    return uid


def class_uid(grade=None, teacher=None, entry=None):
    if grade is None:
        grade = get_grade(entry)
    if teacher is None:
        teacher = get_teacher(entry)
    uid = hashlib.sha1((f"{grade}_{teacher}").encode("utf-8")).hexdigest()
    return uid


def pool_to_student_relations(pool):

    out = {}
    for entry in pool:
        student_name = entry.get("Student")
        # dob = entry.get("Birth Date")
        grade = get_grade(entry)
        teacher = get_teacher(entry)

        phone = entry.get("Phone")

        address1, address2 = get_address12(entry)

        relation = entry.get("Relation")
        # relation_name = entry.get("Parent/Guardian Name") or entry.get("Name")
        relation_name = get_relation_name(entry)
        relation_cell = get_relation_phone(entry)
        relation_email = entry.get("Parent/Guardian Email") or entry.get("Email")

        uid = student_uid(entry)
        if uid not in out:
            out[uid] = {}
        out[uid]["Student"] = student_name
        out[uid]["Grade"] = grade
        out[uid]["Homeroom Teacher"] = teacher
        if "Relations" not in out[uid]:
            out[uid]["Relations"] = []

        relation_info = {
            "Relation": relation,
            "Name": relation_name,
            "Email": format_email(relation_email),
            "Cell Phone": format_phone_link(format_phone(relation_cell)),
        }

        if address1 != withheld_marker:
            relation_info["Address1"] = address1
        if address2 != withheld_marker:
            relation_info["Address2"] = address2
        if phone != withheld_marker:
            relation_info["Phone"] = phone
        out[uid]["Relations"].append(relation_info)

    for uid in out:
        all_relations = out[uid]["Relations"]
        for k in ["Address1", "Address2", "Phone", "Email", "Cell Phone"]:
            all_vals = set([rel.get(k) for rel in all_relations])
            if all_vals == {None}:
                continue
            if len(all_vals) == 1:
                out[uid][k] = all_vals.pop()
                for rel in all_relations:
                    del rel[k]
        out[uid]["Relations"] = all_relations
        if out[uid].get("Cell Phone") and out[uid].get("Phone") == out[uid].get(
            "Cell Phone"
        ):
            del out[uid]["Cell Phone"]
    return out


def linkedHeading(story, text, style):
    # create bookmarkname
    bn = hashlib.sha1((text + style.name).encode("utf-8")).hexdigest()
    # modify paragraph text to include an anchor point with name bn
    h = Paragraph(text + '<a name="%s"/>' % bn, style)
    # store the bookmark name on the flowable so afterFlowable can see this
    h._bookmarkName = bn
    story.append(h)


def url2qrlink(url):
    link = f"<link href='{url}'>{url}</link>"
    return url2qr(link)


def url2qr(url):

    label = "myqr1_" + hashlib.sha1(url.encode("utf-8")).hexdigest()
    out_fn = f"img-{label}-minimal.png"
    if not pathlib.Path(out_fn).is_file():
        # check filesystem and reuse if possible

        qr = qrcode.QRCode(
            version=1,
            error_correction=qrcode.constants.ERROR_CORRECT_L,
            # error_correction=qrcode.constants.ERROR_CORRECT_M,
            # error_correction=qrcode.constants.ERROR_CORRECT_Q,
            # error_correction=qrcode.constants.ERROR_CORRECT_H,
            box_size=40,
            border=1,
        )

        body = url
        qr.add_data(body)
        qr.make(fit=True)

        factory = qrcode.image.pure.PyPNGImage
        img = qr.make_image(
            # fill_color="green",
            # back_color="purple",
            image_factory=factory,
        )
        outfh = open(out_fn, "wb")
        img.save(outfh)
    out = Image(out_fn, 1 * inch, 1 * inch)
    return out


def pool_to_story(pool):

    styles = getSampleStyleSheet()
    # styles.add(ParagraphStyle(name="Justify", alignment=TA_JUSTIFY))

    toch1 = ParagraphStyle(
        name="TOCHeading1",
        fontSize=14,
        leading=16,
        spaceBefore=10,
        spaceAfter=10,
    )
    # tcoh2 = ParagraphStyle(
    #     name="TOCHeading2",
    #     fontSize=12,
    #     leading=18,
    #     spaceBefore=10,
    #     spaceAfter=10,
    # )

    # h1 = ParagraphStyle(
    #     name="Heading1",
    #     fontSize=14,
    #     leading=16,
    #     spaceBefore=10,
    #     spaceAfter=10,
    # )
    h2 = ParagraphStyle(
        name="Heading2",
        fontSize=12,
        leading=18,  # fontName="Helvetica-Bold",
        spaceBefore=10,
        spaceAfter=10,
    )
    h3 = ParagraphStyle(
        # firstLineIndent=30,
        spaceBefore=10,
        # spaceAfter=45,
        name="Heading2",
        fontSize=10,
        leading=22,
        fontName="Helvetica",
    )

    styleSheet = getSampleStyleSheet()

    # body_style = styleSheet["BodyText"]

    teacher_style = ParagraphStyle(
        name="teacher",
        fontSize=14,
        leading=20,  # leftIndent=15
        fontName="Helvetica-Bold",
    )

    teacher_email_style = ParagraphStyle(
        name="teacher_email",
        fontSize=14,
        leading=20,
        leftIndent=15,
        # fontName="Helvetica-Bold",
    )

    details_student_name_style = ParagraphStyle(
        name="studentName", fontSize=12, leading=15, leftIndent=0
    )
    details_class_teacher_style = ParagraphStyle(
        name="details_teacher", fontSize=10, leftIndent=10  # leading=12,
    )
    details_phone_style = ParagraphStyle(
        name="phone",
        fontSize=10,
        # leading=12,
        leftIndent=10,
    )
    details_address_style = ParagraphStyle(
        name="address",
        fontSize=10,
        # leading=12,
        leftIndent=10,
    )

    student_street_style = ParagraphStyle(
        name="studentStreet", fontSize=12, leading=14, leftIndent=20
    )
    student_teacher_style = ParagraphStyle(
        name="studentTeacher", fontSize=12, leading=14, leftIndent=20
    )
    # style = styles["Normal"]
    normal = styles["Normal"]

    centered_title_style = styles["Heading1"]
    centered_title_style.alignment = 1

    centered_subtitle_style = styles["Heading2"]
    centered_subtitle_style.alignment = 1

    centered_style = copy(styles["Normal"])
    centered_style.alignment = 1

    Story = []

    Story.append(PageBreak())
    Story.append(PageBreak())
    Story.append(
        Paragraph("Somerset Elementary School Directory", centered_title_style)
    )

    Story.append(Paragraph("2024-2025", centered_subtitle_style))
    Story.append(Spacer(1, 12))

    Story.append(Paragraph(format_phone_link("240-740-1100"), centered_style))
    Story.append(Paragraph("5811 Warwick Place, Chevy Chase MD 20815", centered_style))
    url1 = "https://www.montgomeryschoolsmd.org/schools/somersetes"
    link1 = f"<link href='{url1}'>{url1}</link>"
    Story.append(Paragraph(link1, centered_style))
    Story.append(url2qr(link1))

    Story.append(Spacer(1, 12))

    Story.append(Spacer(1, 12))
    Story.append(Paragraph("Mr. Travis J Wiebe", centered_style))
    Story.append(Paragraph("Principal", centered_style))
    Story.append(Paragraph(format_email("Travis_J_Wiebe@mcpsmd.org"), centered_style))
    Story.append(Spacer(1, 12))
    Story.append(Spacer(1, 12))
    Story.append(Spacer(1, 12))

    Story.append(Paragraph("Mrs. Bess W Treat", centered_style))
    Story.append(Paragraph("Assistant School Administrator", centered_style))
    Story.append(Paragraph(format_email("Bess_W_Treat@mcpsmd.org"), centered_style))
    Story.append(Spacer(1, 12))
    Story.append(Spacer(1, 12))
    Story.append(Spacer(1, 12))
    # Story.append(Paragraph("Main Office", centered_subtitle_style))

    Story.append(Paragraph("Mrs. Nancy L Conway", centered_style))
    Story.append(Paragraph("School Secretary", centered_style))
    Story.append(Paragraph(format_email("Nancy_L_Conway@mcpsmd.org"), centered_style))
    Story.append(Spacer(1, 12))
    Story.append(Spacer(1, 12))
    Story.append(Spacer(1, 12))

    Story.append(Paragraph("Ms. Susan E Stringham", centered_style))
    Story.append(Paragraph("School Admin Secretary", centered_style))
    Story.append(Paragraph(format_email("Susan_Stringham@mcpsmd.org"), centered_style))
    Story.append(Spacer(1, 12))
    Story.append(Spacer(1, 12))
    Story.append(Spacer(1, 12))

    Story.append(Spacer(1, 12))
    Story.append(Spacer(1, 12))
    Story.append(Spacer(1, 12))
    Story.append(Spacer(1, 12))
    Story.append(Paragraph("Published by the Somerset PTA", centered_style))

    Story.append(PageBreak())

    toc = TableOfContents()
    # toc.levelStyles = [h1]#, h2]
    Story.append(toc)

    style_right = ParagraphStyle(
        name="right", parent=styles["Normal"], alignment=TA_RIGHT
    )

    if True:

        Story.append(PageBreak())

        linkedHeading(Story, "Parent Teacher Association (PTA)", toch1)

        # doing the bare minimum to discourage bots from scraping the whatsapp signup url
        rot13_whatsapp_group_url = "uggcf://pung.jungfncc.pbz/PFDzg9y9Skz8IkW0eDPdn9"
        whatsapp_group_url = codecs.encode(rot13_whatsapp_group_url, "rot_13")
        calendar_url = "https://calendar.google.com/calendar/u/2?cid=Y18xNjY5ZGVlYWNlZmE5ODZiMDAzZDFiNGEwOGE2MzNiOWZiZjM5N2UwNWZjMzZhZTg5MTk0YWVhZTg4OTNmNTI4QGdyb3VwLmNhbGVuZGFyLmdvb2dsZS5jb20"

        # Story.append(Paragraph("PTA", centered_subtitle_style))

        # Story.append(Paragraph(format_email("info@somersetpta.org"), centered_style))
        Story.append(
            KeepTogether(
                [
                    # Paragraph(, h2),
                    Paragraph(
                        """Parent Teacher Association (PTA) The PTA is composed of parent volunteers. All families are welcome at any PTA event or meeting, but only individuals who have joined the PTA and paid annual dues may vote on PTA proposals, budgets, and elect officers. The PTA welcomes all volunteers and any interested board candidates or committee chairs. Elections for officers and board members are generally held in late May or early June. The PTA's mission is to support kids and teachers in their classrooms. We fill an important gap- providing teacher stipends for much-needed school materials, books for classrooms and libraries, tools like microscopes, calculators, as well as hosting before and afterschool activities and enrichment options, and providing help for kids in need, from field trip scholarships to snacks for kids who arrive hungry. The PTA also hosts fun community events, from the Back to School Picnic and the Back to School Classic Race, to the Circle of Giving Dance, and Skate Night. It offers cultural arts assemblies and funds an Adventure Theater enrichment program and performance. Plus, the PTA recognizes and appreciates our teachers and staff throughout the year.""",  # To learn more, visit https://somersetelementary.memberhub.com/.""",
                        normal,
                    ),
                    Spacer(1, 12),
                    Paragraph(format_email("info@somersetpta.org"), centered_style),
                    Spacer(1, 12),
                ]
            )
        )

        Story.append(
            KeepTogether(
                [
                    Paragraph("PTA Calendar", h2),
                    Paragraph(
                        """The PTA maintains an upcoming events calendar at the url shown below, (or scan the QR Code to avoid typing)""",
                        normal,
                    ),
                    Spacer(1, 12),
                    Paragraph(
                        calendar_url,
                        centered_style,
                    ),
                    Spacer(1, 12),
                    url2qr(calendar_url),
                ]
            )
        )
        Story.append(
            KeepTogether(
                [
                    Paragraph("New & International Families WhatsApp Group", h2),
                    Paragraph(
                        """There is a WhatsApp group for new & international families which can be joined via
                        """,
                        normal,
                    ),
                    Spacer(1, 12),
                    Paragraph(whatsapp_group_url, centered_style),
                    Spacer(1, 12),
                    url2qr(whatsapp_group_url),
                ]
            )
        )

    if True:
        Story.append(PageBreak())

        linkedHeading(Story, "Somerset ES School Information", toch1)

        Story.append(
            KeepTogether(
                [
                    Paragraph("New Students", h2),
                    Paragraph(
                        """Families who are new to the neighborhood are encouraged to register as soon as possible. Contact the main office at""",
                        normal,
                    ),
                    Paragraph(format_phone_link("240-740-1100"), centered_style),
                    Paragraph(
                        """Information about enrollment is at
                        http://www.montgomeryschoolsmd.org/info/enroll/index.aspx
                        """,
                        normal,
                    ),
                    url2qr("http://www.montgomeryschoolsmd.org/info/enroll/index.aspx"),
                ]
            )
        )

        Story.append(
            KeepTogether(
                [
                    Paragraph("Back to School Ready", h2),
                    Paragraph(
                        """The following article is a great resource as you begin to prepare your children to return to school.  It broadens the concept of school readiness. I hope that you will take the time to read it and put into practice some of the recommendations.
                        https://www.gettingsmart.com/2019/07/broadening-conceptions-of-back-to-school-readiness/""",
                        normal,
                    ),
                    url2qr(
                        "https://www.gettingsmart.com/2019/07/broadening-conceptions-of-back-to-school-readiness/"
                    ),
                ]
            )
        )

        Story.append(
            KeepTogether(
                [
                    Paragraph("Arrival Information", h2),
                    Paragraph(
                        """For the first day of school, beginning after 8:40 am,  students will meet their teachers outside on the back blacktop and then line up to enter the building. On Tuesday, we will shift to the indoor lineup procedures noted below.""",
                        normal,
                    ),
                    Paragraph(
                        """All students should enter through the front doors and proceed to their designated lineup spot outside their classroom.  Kindergarten students will enter through the south door at the bottom of the bus loop.  Students in grades 1-5 will enter through the main entrance doors at the top of the bus loop. Students purchasing breakfast should go directly to the cafeteria/ APR at 8:40 am.""",
                        normal,
                    ),
                    Paragraph("Bike Riders", h3),
                    Paragraph(
                        """Students who ride their bikes to school are to lock their bikes to the bike rack located at the south end of the building near the purple playground.  Please be sure to provide your child with a bike lock that they can easily use.  It is recommended that bikes be locked at all times.""",
                        normal,
                    ),
                    Paragraph("Scooters and Rollerblades", h3),
                    Paragraph(
                        """If your child uses a scooter to come to school, it must be folded and carried once on school properly.  If your child uses roller blades to get to school, they must remove them prior to entering the school building. Use of a bike, scooter, or rollerblade helmet is mandatory by Maryland law.""",
                        normal,
                    ),
                    Paragraph("Car Riders", h3),
                    Paragraph(
                        """If you are driving your child to school, please drop them off on Warwick at the north driveway entrance to the school (top of the bus loop) so they may use the sidewalk to enter the building through the main entrance. Students should be ready to exit vehicles quickly from the passenger side of your vehicle as our buses need to be able to easily exit the bus loop. """,
                        normal,
                    ),
                    Spacer(1, 12),
                    Paragraph(
                        """The parking lot (entrance on Deal Place) is for staff parking only and will not be accessible for non-staff parking or drop off.  Please do not drop off students at the rear entrance.  All students and families are to use the main entrance on Warwick Place.  There will be a crossing guard located at the intersection of Dorset Avenue and Warwick Place.""",
                        normal,
                    ),
                    Spacer(1, 12),
                    Paragraph(
                        """All students should arrive between 8:40 and 8:50 am to ensure they are ready to enter their classrooms at 8:55 am.  There is no supervision for students prior to 8:40 am and other than those who attend the Bar-T Before-School program, students are not permitted to be in the building prior to 8:40 am, including morning helpers. Please note that attendance is taken at 9:00 AM.  If your child is not in the CLASSROOM at 9:00 AM, your child will be marked absent.""",
                        normal,
                    ),
                ]
            )
        )

        Story.append(
            KeepTogether(
                [
                    Paragraph("Dismissal Information", h2),
                    Paragraph(
                        """Because dismissal is a busy time in the classrooms and main office, we would like to minimize distractions between 3:00 and 3:25 pm.  If you must pick your child up prior to the regular dismissal time of 3:25 pm, please come to the office before 3:00 pm to sign out your child.""",
                        normal,
                    ),
                    Spacer(1, 12),
                    Paragraph(
                        """If your family has an emergency change in dismissal plans, please email your child’s teacher and call the office to ensure your child receives the message as teachers may not be able to check their email during the school day. We are unable to accommodate dismissal changes after 3:00 pm.""",
                        normal,
                    ),
                    Paragraph("Dismissal Schedule:", h2),
                    Paragraph("3:15", h3),
                    Paragraph(
                        """Honor guards, patrols, and office helpers.

    Siblings of Kindergarten walkers and car riders dismissed to the Kindergarten classrooms

    Bar-T students (to the APR):
    Kindergarten and grade 1 (escorted by Bar-T staff)
    Grades 2-5 walk on their own """,
                        normal,
                    ),
                    Paragraph("3:18", h3),
                    Paragraph(
                        """First grade walkers and car riders (south doors by cafeteria)

    Second and third grade walkers and car riders (north doors by main office)

    Kindergarten walkers and car riders are dismissed out their classroom doors""",
                        normal,
                    ),
                    Paragraph("3:20", h3),
                    Paragraph(
                        """Fourth grade walkers and car riders (south doors by cafeteria)

    Fifth grade walkers and car riders (north doors by main office)""",
                        normal,
                    ),
                    Paragraph("3:22", h3),
                    Paragraph(
                        """Kindergarten bus riders (picked up by patrols)

    Buses called by color""",
                        normal,
                    ),
                ]
            )
        )

        Story.append(
            KeepTogether(
                [
                    Paragraph("Bus Information", h2),
                    Paragraph(
                        """Letter from the Bethesda Transportation Depot https://docs.google.com/document/d/1BUH-1MOiz-YLGt_XP3idapAIVefX1wAt/view""",
                        normal,
                    ),
                    url2qr(
                        "https://docs.google.com/document/d/1BUH-1MOiz-YLGt_XP3idapAIVefX1wAt/view"
                    ),
                    Paragraph(
                        """Bus Routes Arrival (AM) https://drive.google.com/file/d/1H6cpsGhsggXCCP7iNXIsvCEam6hSDpXQ/view""",
                        normal,
                    ),
                    url2qr(
                        "https://drive.google.com/file/d/1H6cpsGhsggXCCP7iNXIsvCEam6hSDpXQ/view"
                    ),
                    Paragraph(
                        """Bus Routes Departure (PM) https://drive.google.com/file/d/1yEq_2mEFfJd7f0CoFQvHSu-YMXWr4vEJ/view""",
                        normal,
                    ),
                    url2qr(
                        "https://drive.google.com/file/d/1yEq_2mEFfJd7f0CoFQvHSu-YMXWr4vEJ/view"
                    ),
                    Paragraph(
                        """As a reminder, please arrive early to the bus stops during the first week of school as the drivers work to fine tune the schedule.  Students will be provided with a corresponding colored wristband for their bus to assist them in boarding the correct bus.  Please encourage your child to wear the band for the first two weeks of school. Please also make note of the corresponding route number.  In the event you need to contact the Bethesda Transportation Depot, they refer to buses by route number not color. """,
                        normal,
                    ),
                ]
            )
        )

        bus_table = []
        if bus_table:
            for k, v in [
                ("How do I get home?", "Bracelet color?"),
                ("Bar-T Child Care", "Rainbow"),
                ("Walkers/Car Riders", "Green"),
                ("Red Bus (Route #1114)", "Red"),
                ("Blue Bus (Route #1113)", "Blue"),
                ("Orange Bus (Route #1127)", "Orange"),
                ("Walkers/Car Riders", "Green"),
            ]:
                bus_table.append(
                    [
                        Paragraph(k),
                        Paragraph(v, style_right),
                    ]
                )
        if bus_table:
            t = Table(
                bus_table,
                style=[
                    ("LINEBELOW", (-2, -1), (-1, -1), 0.25, colors.black),
                    # ("GRID", (0, 0), (-1, -1), 0.5, colors.black),
                ],
            )

            Story.append(KeepTogether(t))

        Story.append(
            KeepTogether(
                [
                    Paragraph("Before/After School Care", h2),
                    Paragraph(
                        """Please see the linked information below regarding BAR-T, Somerset’s before and after school child care provider. https://www.bar-t.com/program/kids-club/""",
                        normal,
                    ),
                    url2qr("https://www.bar-t.com/program/kids-club/"),
                ]
            )
        )

        Story.append(
            KeepTogether(
                [
                    Paragraph("New Staff", h2),
                    Paragraph(
                        """We are pleased to welcome the following staff members to Somerset!""",
                        normal,
                    ),
                    Paragraph(
                        """Barbara Davis\nPart-Time Special Education Teacher\nAshburton ES""",
                        normal,
                    ),
                    Paragraph(
                        """Ladraine Greene\nPart-Time Special Education Teacher\nNorth Carolina""",
                        normal,
                    ),
                    Paragraph(
                        """Kate Waldmann\nSpecial Education Paraeducator\nDCPS""",
                        normal,
                    ),
                ]
            )
        )

        Story.append(
            KeepTogether(
                [
                    Paragraph("Important Dates", h2),
                    Paragraph("""August 23 - Sneak a Peek Day""", normal),
                    Paragraph(
                        """    August 25 - PTA PlayDate for New Families - Norwood Park (10:00 am)""",
                        normal,
                    ),
                    Paragraph("""    August 26  - First Day of School""", normal),
                    Paragraph(
                        """    September 2 - Labor Day Holiday - Schools and Offices Closed""",
                        normal,
                    ),
                    Paragraph(
                        """    September 6 - PTA Back to School Dinner and a Movie""",
                        normal,
                    ),
                    Paragraph(
                        """    September 10 - Back to School Night (6:30-8:00 PM) for all grades""",
                        normal,
                    ),
                    Paragraph("""    September 20 - International Night""", normal),
                    Paragraph(
                        """    September 27 - Early Release Day (Students dismissed at 12:55 pm)""",
                        normal,
                    ),
                    Paragraph(
                        """    October 3 - No School for Students or Teachers""", normal
                    ),
                    Paragraph(
                        """    October 6 - PTA Back to School Classic Race""", normal
                    ),
                    Paragraph("""    October 9 - Student Portraits""", normal),
                    Paragraph("""    October 18 - No School for Students""", normal),
                    Paragraph("""    November 1 -  Fall Festival""", normal),
                ]
            )
        )

        Story.append(
            KeepTogether(
                [
                    Paragraph("MCPS Calendar", h2),
                    Paragraph(
                        """The MCPS Comprehensive School Calendar for the 2024-2025 school year can be found at
                        https://ww2.montgomeryschoolsmd.org/calendar/""",
                        normal,
                    ),
                    url2qr("https://ww2.montgomeryschoolsmd.org/calendar/"),
                ]
            )
        )

        Story.append(
            KeepTogether(
                [
                    Paragraph("Health Information", h2),
                    Paragraph(
                        """A copy of the Maryland State Immunization Guidelines for the 2024-25 school year is available """,
                        normal,
                    ),
                    url2qr(
                        "https://health.maryland.gov/phpa/OIDEOR/IMMUN/Shared%20Documents/2024-2025_School_IZ_%20Requirements_Final.pdf"
                    ),
                    Paragraph(
                        """New students should present proper documentation of required immunizations at the time of enrollment. If a family does not provide documentation for the required immunizations but presents evidence of an appointment to obtain documentation or immunization records within 20 calendar days of the date of enrollment, the student may be enrolled. However, if the documentation is not provided immediately following the scheduled date, the student must be excluded from school and marked absent until proof of immunization is received. If you have questions or concerns about these requirements, please contact our School Health Technician at 240-740-1102, on or after August 20, 2024.""",
                        normal,
                    ),
                    Spacer(1, 12),
                    Paragraph(
                        """To ensure a smooth, lice free opening of school, we are requesting that families inspect their children’s heads prior to their return to school and every few weeks thereafter.  Should lice be evident, please notify our School Health Technician.  She will be able to provide you with information regarding treatment of a lice infestation.  In addition, upon your child’s return to school following an episode of lice, our School Health Technician will need to inspect the student’s head to assure the absence of live lice and nits located with ½ inch of the scalp.  If the nits are located ½ inch or more from the scalp, a student may attend school.""",
                        normal,
                    ),
                    Spacer(1, 12),
                    Paragraph(
                        """Please be sure that you have provided us with up to date contact information including phone numbers and email addresses so that the health room or school can reach you directly.  Be sure to update information using ParentVue.""",
                        normal,
                    ),
                ]
            )
        )

        Story.append(
            KeepTogether(
                [
                    Paragraph("Attendance", h2),
                    Paragraph(
                        """Please refer to MCPS Regulation JEA-RA Student Attendance https://ww2.montgomeryschoolsmd.org/departments/policy/pdf/jea-ra%20nondiscrimination.pdf """,
                        normal,
                    ),
                    url2qr(
                        "https://ww2.montgomeryschoolsmd.org/departments/policy/pdf/jea-ra%20nondiscrimination.pdf"
                    ),
                    Paragraph(
                        """for additional information regarding student attendance and please remember the importance of regular attendance in school.  Please try to schedule your vacations or family activities when school is not in session. We realize that sometimes vacation opportunities come up unexpectedly; however, these absences are considered unlawful and the absence(s) will be recorded as unexcused.  Students learn when they are in school.  Please support us by making sure your child comes to school on time each day.  If your child is late, please sign them in at the main office.""",
                        normal,
                    ),
                ]
            )
        )

        Story.append(
            KeepTogether(
                [
                    Paragraph("Lunch Information", h2),
                    Paragraph(
                        """Meal prices: Breakfast—$1.30; Lunch—$2.55 (Elementary), $2.80 (Middle and High) """,
                        normal,
                    ),
                    Paragraph(
                        """Students in Maryland who qualify for reduced-price meals will not be charged for breakfast or lunch. FREE AND REDUCED-PRICE MEAL APPLICATIONS Apply online NOW so that benefits will be in place when school starts! Families who meet certain federal income standards are eligible for free or reduced-price meal benefits. All children use their MCPS ID number at the register so confidentiality is maintained and no child is overtly identified as receiving free or reduced-price meals. Only one application is needed for all students in a household. A new application must be completed for any family requesting assistance this school year. The online application is now open for families to submit applications online at  http://www.MySchoolApps.com
    """,
                        normal,
                    ),
                    url2qr("http://www.myschoolapps.com"),
                    Paragraph(
                        """For more information and How-To videos, visit the Food Services website http://www.montgomeryschools.org/departments/food-and-nutrition/meal-payments/#FARMS""",
                        normal,
                    ),
                    url2qr(
                        "http://www.montgomeryschools.org/departments/food-and-nutrition/meal-payments/#FARMS"
                    ),
                ]
            )
        )

        Story.append(
            KeepTogether(
                [
                    Paragraph("School Supplies", h2),
                    Paragraph(
                        """The school supply list is available at  https://drive.google.com/file/d/1_kWkA8yM8vTPURe1rwSCHLMmrznQ6HLe/view""",
                        normal,
                    ),
                    url2qr(
                        "https://drive.google.com/file/d/1_kWkA8yM8vTPURe1rwSCHLMmrznQ6HLe/view"
                    ),
                ]
            )
        )

        Story.append(
            KeepTogether(
                [
                    Paragraph("Staff Directory", h2),
                    Paragraph(
                        """The Somerset staff directory is available at https://www.montgomeryschoolsmd.org/schools/somersetes/staff/directory/""",
                        normal,
                    ),
                    url2qr(
                        "https://www.montgomeryschoolsmd.org/schools/somersetes/staff/directory/"
                    ),
                ]
            )
        )

        Story.append(
            KeepTogether(
                [
                    Paragraph("For Kindergarten Only", h2),
                    Paragraph(
                        """On the first day of school, kindergarten students should line up outside on the blacktop near their teacher.  Staff will be available to escort kindergarten bus riders to the back of the school. Each teacher will be holding a placard with their name on it, so you can find them easily.  If you are accompanying your child to school, please say goodbye outside and allow your child to enter the building with their teachers and classmates.  We will have staff readily available to escort the children to the correct location.  This makes the transition for children so much easier.  We promise to take the best care of them on the first day and every day!""",
                        normal,
                    ),
                    Spacer(1, 12),
                    Paragraph(
                        """It is important to note that the kindergarten year is a very special one during which many of the attitudes toward school and learning begin to take shape.  In addition to learning the kindergarten objectives, we want all children to have a positive perception about school and themselves.  The kindergarten instructional program deals with child development in the areas of language and thinking skills, concepts in math, science, social studies, art, and music, as well as social, emotional, and physical development.""",
                        normal,
                    ),
                    Spacer(1, 12),
                    Paragraph(
                        """Students who ride the bus will be picked up at the stop located closest to your home.  A copy of the bus schedule is included above.  Kindergarten students who are bus riders will ride the bus at the beginning and end of the school day.  Please ensure you are early to the bus stop for the first two weeks of school as the buses work to normalize pick up and drop off. Kindergarten students will not be dropped off at the end of the day unless an adult is waiting to pick them up. """,
                        normal,
                    ),
                ]
            )
        )

    Story.append(PageBreak())
    ptext = "FAQ"
    linkedHeading(Story, ptext, toch1)

    Story.append(Spacer(1, 12))

    # SOMERSET A to Z

    absence_url = "https://www.montgomeryschoolsmd.org/schools/somersetes/absence-form/"
    Story.append(
        KeepTogether(
            [
                Paragraph("Absences", h2),
                Paragraph(
                    """If a student is going to be absent for any reason, parents are asked to inform the school. You can report an absence via""",
                    normal,
                ),
                Paragraph(absence_url, centered_style),
                url2qr(absence_url),
                Paragraph(
                    """You can also telephone the school office prior to 9 am at""",
                    normal,
                ),
                Paragraph(
                    f"Phone: {format_phone_link(format_phone('240-740-1100'))}",
                    details_phone_style,
                ),
                Spacer(1, 12),
                Paragraph(
                    """After missing five consecutive days of school, it's requested that you submit a doctor's note. Maryland State Department of Education rules for attendance: A student is counted present for a full day if a student is in school for four hours or more of the school day. A student is counted as absent for a half day if he or she arrives more than two hours after the start of the school day, leaves more than two hours before the end of the school day or leaves school for more than two hours during the day. A student is considered tardy if he or she arrives after the last bell and within the first two hours of the school day.""",
                    normal,
                ),
            ]
        )
    )

    Story.append(
        KeepTogether(
            [
                Paragraph(
                    "Advertising",
                    h2,
                ),
                Paragraph(
                    """Permission to distribute advertising material of any kind, in the school or on the grounds, must follow the guidelines set forth by the Board of Education. Please consult the school office staff before distributing materials.""",
                    normal,
                ),
            ]
        )
    )

    # (see also Departure from School)
    Story.append(
        KeepTogether(
            [
                Paragraph("Back-to-School Classic 8k/2k", h2),
                Paragraph(
                    """The Back-to-School Classic has been run for over 30 years. Somerset parents and students work together to host a certified 8K road race, a 2K run/walk, as well as special student fun runs. This event attracts hundreds of runners from the metro area and is one of the PTA's largest fundraising activities.""",
                    normal,
                ),
            ]
        )
    )

    Story.append(
        KeepTogether(
            [
                Paragraph("Back-to-School Night", h2),
                Paragraph(
                    """In September, parents are invited to meet the teachers and visit their student's classrooms for an explanation of the school year curriculum and classroom policies and practices. There are separate nights for grades K-2 and 3-5.""",
                    normal,
                ),
            ]
        )
    )

    Story.append(
        KeepTogether(
            [
                Paragraph("Bar-T Kids Club", h2),
                Paragraph(
                    """Before and after school programs offers a place for students to learn, play and enjoy the supportive Bar-T community. For information regarding before and after school child care ChildCare please call: Bar-T Kids Club at 240-364-4196 or vist https://www.bar-t.com/program/kids-club/""",
                    normal,
                ),
                url2qr("https://www.bar-t.com/program/kids-club/"),
            ]
        )
    )

    Story.append(
        KeepTogether(
            [
                Paragraph("Bicycles and Scooters", h2),
                Paragraph(
                    """Students are permitted to ride bikes and scooters to school and are required to wear bike helmets. All bikes and scooters must be parked and locked at the bike rack located on the south side of the building. Bikes and scooters are not permitted anywhere else on school grounds. MCPS and Somerset Elementary does not assume responsibility for bicycles and scooters brought to school.""",
                    normal,
                ),
            ]
        )
    )

    # "Birthday or Other Personal Celebrations"
    # """Party invitations for celebrations must be sent to students at their home addresses and may not be distributed at school. With the approval of classroom teachers, limited school celebration, such as store bought cookies or cupcakes to share with classmates, is usually permitted. Please contact the classroom teacher directly to discuss his/her policy."""

    Story.append(
        KeepTogether(
            [
                Paragraph("Bus Transportation", h2),
                Paragraph(
                    """MCPS provides bus service for Somerset students who live outside of the walking boundaries. For questions about the bus service, please call William Stapleton at 301-469-1068. For route stops and schedules, visit the MCPS website and choose Students, then Bus Transportation. Select Bus Routes by School, then Somerset ES.""",
                    normal,
                ),
            ]
        )
    )

    Story.append(
        KeepTogether(
            [
                Paragraph("Camp Summerset", h2),
                Paragraph(
                    """The PTA runs a summer camp designed for pre-K (age 4) - 5th grade (rising 6th graders) that offers an array of activities including arts, sports, yoga, games, swimming, and trips into the city. Camp starts the week after school is out and is staffed by Somerset teachers and other professionals. Information and registration are available in January at www.campsummerset.com.""",
                    normal,
                ),
            ]
        )
    )

    if False:
        Story.append(
            KeepTogether(
                [
                    Paragraph("Career Day", h2),
                    Paragraph(
                        """Individuals, many of them Somerset parents, representing a variety of professions and skills, visit Somerset to talk the about different career paths. This event is usually held in late April / early May.""",
                        normal,
                    ),
                ]
            )
        )

    Story.append(
        KeepTogether(
            [
                Paragraph("Change of Address and Telephone Numbers", h2),
                Paragraph(
                    """The school must have current addresses and phone numbers (home, cell and work) for all parents and guardians. Please remember to let the school office know immediately of any changes to contact information on your emergency cards.""",
                    normal,
                ),
            ]
        )
    )

    # Character Education
    # In 1997, Somerset became a Community of Caring school by adopting a character education program that acknowledges the importance of our student's social and emotional development, along with their academic achievements. Somerset hopes to foster a learning environment that embodies the following values: Trust, Caring, Respect, Responsibility, and Family. In 1998, the Maryland Board of Education mandated that each public school in the state implement a character education program toward the goal of integrating the concepts of such programs into school curricula and activities. All schools in the Bethesda-Chevy Chase (B-CC) Cluster have since agreed to adopt the Community of Caring as their vehicle for character education. This means that from kindergarten through twelfth grade our students' character education will be based on a common philosophy and vocabulary. Long before Community of Caring, Somerset worked to integrate service projects into the curriculum, Service Learning began in 1989 under the name "SKIP" (Somerset Kids Participating). In 1998, Service Learning projects were incorporated within the broader goals of the Community of Caring. Somerset now has Service Learning projects integrated into the curriculum at all grade levels.

    # Communications
    # Most bulletins and notices go home on Tuesdays or Fridays.
    # Each grade also has a page or series of pages on the school website where teachers communicate with parents and families through class or grade newsletters, as well as share learning resources and photos of activities at school. The PTA sends out a weekly email (the TIN) on Sundays.

    Story.append(
        KeepTogether(
            [
                Paragraph("Cultural Arts", h2),
                Paragraph(
                    """As part of the school program, Somerset students have many opportunities to extend their appreciation for the cultural arts. With support from PTA funds, professional drama, dance, and music groups entertain Somersets students at many in-school performances. Artist-in-residence programs help integrate the arts into the curriculum in individual grades. The school organizes many field trips to support the instructional program. Students may attend performances at the Kennedy Center, Strathmore, and also visit theaters, museums, and other cultural centers.""",
                    normal,
                ),
            ]
        )
    )

    Story.append(
        KeepTogether(
            [
                Paragraph("Delayed Opening and Emergency Closing", h2),
                Paragraph(
                    """In the event of severe weather, Somerset may delay opening by 2 hours, close 2-1/2 hours early, or close for the day. If school is closed early due to weather conditions, all students will be sent home according to the emergency instructions provided by parents/guardians. In addition, any after school and evening activities scheduled at the school will be canceled.""",
                    normal,
                ),
                Paragraph(
                    """PLEASE NOTE: If severe weather conditions exist anywhere in the general area, please stay informed. MCPS covers a large and diverse weather region, and may declare a school closure even if severe conditions do not exist in Somerset. Call the MCPS Emergency Hot-Line at 301-279-3673 for recorded emergency information, or check the MCPS website at www.montgomeryschoolsmd.org. Subscribe to MCPS QuickNotes for weather-related email messages.""",
                    normal,
                ),
                Paragraph(
                    """You may also sign up for SMS text and email messages with AlertMCPS. Twitter updates can be accessed at twitter.com/mcps. The school will also post information on somerset-net listserv. For delayed opening or emergency closing information for Bar-T at Somerset, call Bar-T Kids Club at 240-364-4196.""",
                    normal,
                ),
            ]
        )
    )

    Story.append(
        KeepTogether(
            [
                Paragraph("Departure from School", h2),
                Paragraph(
                    """Students are dismissed at 3:18 pm, and should leave school promptly at that time. Students enrolled in the Bar-T Before or After School programs, or attending a Somerset academic club or activity, will meet at their designated locations. If a student is to be picked up by someone other than their parent or guardian, the school must have wriften authorization from the parent or guardian.""",
                    normal,
                ),
            ]
        )
    )

    Story.append(
        KeepTogether(
            [
                Paragraph("Directory", h2),
                Paragraph(
                    """This directory is available in both printed and electronic forms. The electronic form is a PDF with hyperlinks between sections, with several helpful features. When looking at a student's details, clicking on the street name, will take you to all other students on that street, or clicking on the teacher's name will take you to all other students in that class. Contact info@somersetpta.org for more information or to request an up to date PDF.""",
                    normal,
                ),
            ]
        )
    )
    Story.append(
        KeepTogether(
            [
                Paragraph("Discipline", h2),
                Paragraph(
                    """Copies of Somerset's discipline policy are available in the school office or at https://www.montgomeryschoolsmd.org/schools/somersetes/about/""",
                    normal,
                ),
                url2qr("https://www.montgomeryschoolsmd.org/schools/somersetes/about/"),
            ]
        )
    )

    Story.append(
        KeepTogether(
            [
                Paragraph("Dogs", h2),
                Paragraph(
                    """A county ordinance prohibits dogs from being on school grounds while school is open. If you walk to school with your dog, please do not bring him or her on school grounds. If you decide to walk your dog on school grounds outside of school hours, please be courteous and pick up after your pet at all times. Dogs are never permitted to be on the turf field, even outside of school hours.""",
                    normal,
                ),
            ]
        )
    )

    Story.append(
        KeepTogether(
            [
                Paragraph("Early Dismissal", h2),
                Paragraph(
                    """A student who needs to be dismissed early for any reason must bring an explanatory note. No student will be dismissed during school hours to anyone other than his or her parent (s) without written permission. Students are to be picked up at the office. A parent or guardian must provide requested information on the sign-out log located on the office counter.""",
                    normal,
                ),
            ]
        )
    )

    Story.append(
        KeepTogether(
            [
                Paragraph("Early Release Day: Half Days for All Students", h2),
                Paragraph(
                    """There are several half-days during the year when all students are dismissed at 12:55 pm (see School Calendar). Many are Teacher Professional Days, when the Board of Education holds workshops for teachers. Teacher conferences with parents are also scheduled on half-days. Bar-T Kids Club is open to students enrolled in the After-School program.""",
                    normal,
                ),
            ]
        )
    )

    Story.append(
        KeepTogether(
            [
                Paragraph("Enrollment Entrance Requirements", h2),
                Paragraph(
                    """Students who are five years old before September 1st are eligible to attend kindergarten that year. In order to register your student, you must present a completed registration form, birth certificate or passport, current rental lease, property tax bill or utility bill, and a completed immunization and health inventory. Detailed information can be found on the MCPS website under Getting Started in the Parents section.""",
                    normal,
                ),
            ]
        )
    )

    Story.append(
        KeepTogether(
            [
                Paragraph("ELD (English Language Development)", h2),
                Paragraph(
                    """Somerset holds special classes during the school day for students who cannot understand, speak, read or write English well enough to follow regular classroom instruction. This special help continues until the student knows enough of the language to learn within the regular classroom.""",
                    normal,
                ),
            ]
        )
    )

    Story.append(
        KeepTogether(
            [
                Paragraph("Field", h2),
                Paragraph(
                    """The field at Somerset Elementary is artificial turf with an organic infill. It is used by Somerset students only during school hours and is open to the public after school hours. Please do not take/use food, sunflower seeds, tobacco products, or gum on the field. No metal cleats are allowed or any other devices that might rip or puncture the turf. No dogs or pets are allowed on at any time.""",
                    normal,
                ),
            ]
        )
    )

    Story.append(
        KeepTogether(
            [
                Paragraph("Field Day", h2),
                Paragraph(
                    """
                    In late spring, grades K-5 have a full day of games and sports on the Somerset field. It is organized by the physical education teacher with the help of other staff members and parent volunteers. """,
                    normal,
                ),
            ]
        )
    )

    Story.append(
        KeepTogether(
            [
                Paragraph("Financial Help", h2),
                Paragraph(
                    """Families with limited incomes may apply to the Board of Education for free or reduced cost breakfasts and lunches. An application form is sent home with all students at the start of each year. You also can apply in confidence to the Principal for help towards the cost of field trips. Limited scholarships also are available for After School Program classes. No student need miss class outings because of a limited family budget. Families can apply on at https://www.MySchoolApps.com""",
                    normal,
                ),
                url2qr("https://www.MySchoolApps.com/"),
            ]
        )
    )

    Story.append(
        KeepTogether(
            [
                Paragraph("Foundation", h2),
                Paragraph(
                    """The Somerset Elementary School Educational Foundation is a non-profit, charitable organization created by committed parents, staff and community leaders to enrich the learning experience of our children at Somerset, and to reach out to the wider community. The Foundation solicits funds from Somerset ES families to support initiatives that improve the educational resources available to students in a way that is consistent with the policies of Montgomery County Public Schools. The Foundation works closely with the school's administration and the PTA to identify priorities. To learn more about the foundation or to get involved, please visit somerset-foundation.org.""",
                    normal,
                ),
            ]
        )
    )

    Story.append(
        KeepTogether(
            [
                Paragraph("Guidance Counselor", h2),
                Paragraph(
                    """Somerset's counselor works with school staff, helping students resolve problems or concerns that may affect their school performance. Students may refer themselves or be referred by teachers or parents. They see the counselor individually or in small groups. The counselor also works with an entire class to address problems such as teasing, fighting, and other social difficulties.""",
                    normal,
                ),
            ]
        )
    )

    # "Halloween Parade"
    # """Halloween is celebrated with gusto at Somerset, at the end of October. On this occasion the students (and some teachers and parents) don costumes at school and hold a grand march around the neighborhood in the afternoon. Each classroom then holds a Halloween-themed celebration."""

    Story.append(
        KeepTogether(
            [
                Paragraph("Highly Gifted", h2),
                Paragraph(
                    """Also known as The Elementary Center Programs for the Highly Gifted program. Through testing, observation, and other methods, the school identifies gifted students and provides appropriate alternative school programs. Teachers may challenge students by presenting work at higher grade levels or by creating ability groupings. The school also screens students to determine their eligibility for special countywide programs. Countywide testing for gifted/talented designation occurs in 2nd grade. Visit the MCPS website under Students, and Special Programs.""",
                    normal,
                ),
            ]
        )
    )

    Story.append(
        KeepTogether(
            [
                Paragraph("Illness and Medication", h2),
                Paragraph(
                    """A school nurse or health technician works in Somerset's Health Room five days a week. Should a student be injured or become ill during the school day, the office will notify the parents at once. The student will rest until parents arrive. If neither parent can be reached, the school will call the alternate person specified on the students enrollment card.""",
                    normal,
                ),
                Paragraph(
                    """The school is not authorized to give any medicine without a form signed by a doctor. These forms are available in the school office. MCPS policy requires all medication must be delivered to school by an adult and in the original container. School-held medicines are only available during regular school hours of 9:00 am - 3:00 pm.""",
                    normal,
                ),
            ]
        )
    )

    Story.append(
        KeepTogether(
            [
                Paragraph("Instrumental Music", h2),
                Paragraph(
                    """Students in grades 4 and 5 are eligible to learn and perform instrumental music during the school day. Instruction is free, but students must provide instruments. Rentals are available from local music stores. Limited scholarships are available for rentals for qualified students.""",
                    normal,
                ),
            ]
        )
    )

    # "Internet - also see Web Sites"
    # """The Montgomery County Public School system website is www.montgomeryschoolsmd.org the Somerset Elementary School website is www.montgomeryschoolsmd.org/schools/somersetes and the Somerset PTA website is located at www.somersetpta.org."""

    Story.append(
        KeepTogether(
            [
                Paragraph("Kindergarten Orientation", h2),
                Paragraph(
                    """An open house for prospective kindergarten students is held in early spring. Parents may register their students for school and learn more about the school and curriculum while the students visit kindergarten classes and meet their prospective teachers.""",
                    normal,
                ),
            ]
        )
    )

    Story.append(
        KeepTogether(
            [
                Paragraph("Lice", h2),
                Paragraph(
                    """MCPS has adopted a No-Nit policy. This means that a student will be sent home from school if lice or lice eggs (nits) are detected on the hair or scalp. The student will be re-admitted to school only after treatment has been administered visible eggs have been removed. The PTA is committed to educating parents about the lice policy and helping families whose students have lice.""",
                    normal,
                ),
            ]
        )
    )

    Story.append(
        KeepTogether(
            [
                Paragraph("Lost and Found", h2),
                Paragraph(
                    """All items found in the building and on the grounds are placed in the Lost and Found. Lost and found is located on the ground floor outside the All Purpose Room. Please reclaim any lost items as soon as possible. Any unclaimed items are given to charitable organizations at the end of each term. Identification is easier if all personal items are clearly marked. Contact the school office for the Lost and Found location. Valuables such as jewelry or watches are kept in the office.""",
                    normal,
                ),
            ]
        )
    )

    Story.append(
        KeepTogether(
            [
                Paragraph("Meals", h2),
                Paragraph(
                    """Students can bring lunch from home or purchase meals at the cafeteria. Payment is made in exact change or through a myschoolbucks.com lunch account plan. Each student is given a student ID number, which is keyed in each time the student purchases food from the cafeteria. Notices are sent home when a student's account balance is low. Any funds in a student's account at the end of the school year are carried over to the next school year. No refunds are given. myschoolbucks.com also allows you to monitor the purchases your child makes and allows you to block specified items from being purchased. To set up a lunch account visit https://www.myschoolbucks.com""",
                    normal,
                ),
                url2qr("https://www.myschoolbucks.com"),
            ]
        )
    )

    Story.append(
        KeepTogether(
            [
                Paragraph("Media Center", h2),
                Paragraph(
                    """Somerset's Media Center supports the instructional needs of the students and staff and provides an environment that promotes an appreciation for literature and reading. The Media Center has more than 8,500 print and non-print resources including books, magazines, CD-ROMs, and videos. It operates on an open and flexible schedule. Students may come individually with a pass, in small groups to do research, or in whole classes for research and instruction in information seeking strategies.""",
                    normal,
                ),
            ]
        )
    )

    if False:
        Story.append(
            KeepTogether(
                [
                    Paragraph("DEI/NAACP Parents' Council Representative", h2),
                    Paragraph(
                        """The Parents' Council of the National Association for the Advancement of Colored People seeks to empower parents and guardians of African-American and other minority students enrolled in MCPS who share the goal of equal education for all students. The Parents' Council is composed of representatives from each school. The Council meets monthly throughout the calendar year to share information that parents can use to enhance their student's chances of success. The phone number for the Council's office is 301-657-2062. Somerset's NAACP Rep also serves as a member of the PTA's Board of Directors.""",
                        normal,
                    ),
                ]
            )
        )

    Story.append(
        KeepTogether(
            [
                Paragraph("New and International Families", h2),
                Paragraph(
                    """The PTA hosts a number of events for New and International families. Please see the PTA Web site for a list of events.www.somersetpta.com . Open House Held on Columbus Day every year, this event gives parents a chance to visit classes from 9:00 am to 11:30 pm to see their student's classroom in action.""",
                    normal,
                ),
            ]
        )
    )

    if False:
        Story.append(
            KeepTogether(
                [
                    Paragraph("Parent Teacher Association (PTA)", h2),
                    Paragraph(
                        """Parent Teacher Association (PTA) The PTA is composed of parent volunteers. All families are welcome at any PTA event or meeting, but only individuals who have joined the PTA and paid annual dues may vote on PTA proposals, budgets, and elect officers. The PTA welcomes all volunteers and any interested board candidates or committee chairs. Elections for officers and board members are generally held in late May or early June. The PTA's mission is to support kids and teachers in their classrooms. We fill an important gap- providing teacher stipends for much-needed school materials, books for classrooms and libraries, tools like microscopes, calculators, as well as hosting before and afterschool activities and enrichment options, and providing help for kids in need, from field trip scholarships to snacks for kids who arrive hungry. The PTA also hosts fun community events, from the Back to School Picnic and the Back to School Classic Race, to the Circle of Giving Dance, and Skate Night. It offers cultural arts assemblies and funds an Adventure Theater enrichment program and performance. Plus, the PTA recognizes and appreciates our teachers and staff throughout the year. To learn more, visit https://somersetelementary.memberhub.com/.""",
                        normal,
                    ),
                    url2qr("https://somersetelementary.memberhub.com/"),
                    #                Paragraph(
                    #                    """The PTA also hosts fun community events, from the Back to School Picnic and the Back to School Classic Race, to the Rock 'N Roll Circle of Giving Dance, and Skate Night. It offers cultural arts assemblies and funds a playwright in residence for the fifth grade. Plus, the PTA recognizes and appreciates our teachers and staff throughout the year. To learn more, visit www.somersetpta.org.""",
                    #                    normal,
                    #                ),
                ]
            )
        )

    Story.append(
        KeepTogether(
            [
                Paragraph("Playdates", h2),
                Paragraph(
                    """Arrangements for play dates must be made at home. If your student rides a school bus, a note should be given to their teacher indicating the change in dismissal arrangements.""",
                    normal,
                ),
            ]
        )
    )

    Story.append(
        KeepTogether(
            [
                Paragraph("Playground", h2),
                Paragraph(
                    """Students may play on school grounds only with adult supervision. """,
                    normal,
                ),
            ]
        )
    )

    Story.append(
        KeepTogether(
            [
                Paragraph("Report Cards", h2),
                Paragraph(
                    """Kindergartners receive a checklist report in January and June.""",
                    normal,
                ),
                Paragraph(
                    """Grades 1 - 5 receive report cards in November, February, April and June. Teachers may send written notices or make calls regarding possible low scores to parents by the end of the sixth week in each nine-week grading period.""",
                    normal,
                ),
            ]
        )
    )

    Story.append(
        KeepTogether(
            [
                Paragraph("Retention", h2),
                Paragraph(
                    """A parent conference will be scheduled at least a month before the end of the school year if there is a possibility that a student cannot be promoted.""",
                    normal,
                ),
            ]
        )
    )

    Story.append(
        KeepTogether(
            [
                Paragraph("Room Parents", h2),
                Paragraph(
                    """Room parents work with the teacher to identify specific tasks that need to be performed throughout the year and then recruit other parents to help with these tasks. Examples include arranging class parties, chaperoning field trips, and helping with classroom and PTA projects. They also provide email communication for parents in the classroom. In addition to providing classroom help, room parents act as Community of Caring representatives and first contacts for new families. To volunteer, contact the Room Parent Coordinator.""",
                    normal,
                ),
            ]
        )
    )

    # "School Closing: Emergency - see Delayed Opening and Emergency Closing"

    Story.append(
        KeepTogether(
            [
                Paragraph("Movie Night", h2),
                Paragraph(
                    """Movie night is held in September at the Somerset Elementary Field. Families may enjoy pizza dinner before watching a movie with other families. Teachers, students, siblings and families all enjoy the relaxed time together to celebrate the kick-off of the school year.""",
                    normal,
                ),
            ]
        )
    )

    Story.append(
        KeepTogether(
            [
                Paragraph("Student Portraits/School Photographs", h2),
                Paragraph(
                    """A professional photographer takes individual photos of students in the Fall. Class and individual photos are taken in the Spring.""",
                    normal,
                ),
            ]
        )
    )

    Story.append(
        KeepTogether(
            [
                Paragraph("International Night", h2),
                Paragraph(
                    """This evening is a chance to share the wealth of cultural diversity that our students and their families bring to Somerset. There are numerous exhibits displaying the homelands of or places of interest to Somerset students, international foods to sample, and musical and cultural performances from around the world, this event is organized by the Somerset PTA.""",
                    normal,
                ),
            ]
        )
    )

    Story.append(
        KeepTogether(
            [
                Paragraph("Sneak Peek", h2),
                Paragraph(
                    "Traditionally held the weekday preceding the start of school, this event allows students to visit their classrooms and meet their new teachers."
                    "",
                    normal,
                ),
            ]
        )
    )

    if False:
        Story.append(
            KeepTogether(
                [
                    Paragraph("Somerset Organized Service (S.O.S.)", h2),
                    Paragraph(
                        """The Somerset Organized Service or S.O.S. is a service program for 5th graders. At the end of 4th grade, students are offered an opportunity to fill out an application listing their top priorities for service positions. The choices include announcers, greeters/assembly assistants, honor guard, kindergarten patrols, office assistants, safety patrols, and ambassadors. The students are selected for one of their priority choices. Through the program, 5th graders enhance their leadership and responsibility skills.""",
                        normal,
                    ),
                ]
            )
        )

    Story.append(
        KeepTogether(
            [
                Paragraph("Special Needs", h2),
                Paragraph(
                    """Parents with questions regarding special education issues should contact Special Education Teachers, or the chair of the PTA Special Needs Committee.""",
                    normal,
                ),
            ]
        )
    )

    Story.append(
        KeepTogether(
            [
                Paragraph("Staff Appreciation", h2),
                Paragraph(
                    """The commitment and excellence of our teachers and staff is the key to making Somerset an outstanding school. The PTA recognizes the wonderful and important jobs of these professionals in several ways during the course of the year.""",
                    normal,
                ),
            ]
        )
    )

    # Story.append(
    #     KeepTogether(
    #         [
    #             Paragraph("Student Government Association (SGA)", h2),
    #             Paragraph(
    #                 """Somerset's student council consists of an elected president, vice-president, secretary, treasurer, and two representatives (one boy, one girl) from each class in grades 2 to 5. Students in those grades vote in the Fall after a lively election campaign that lasts for a week. The SGA council meets during school hours to discuss student concerns and ideas. The SGA also organizes school activities and collection drives to benefit student's charities.""",
    #                 normal,
    #             ),
    #         ]
    #     )
    # )

    Story.append(
        KeepTogether(
            [
                Paragraph("Suspension", h2),
                Paragraph(
                    """As a last resort, the principal may suspend a student for up to ten days in cases of extreme misbehavior. A student with any kind of weapon must be suspended. This is a mandatory MCPS policy.""",
                    normal,
                ),
            ]
        )
    )

    Story.append(
        KeepTogether(
            [
                Paragraph("Tardiness", h2),
                Paragraph(
                    """School begins at 9 am. All students arriving after 9 am will be considered tardy. A note of explanation and an adult should accompany the student when he or she arrives. All students who arrive after the 9 am bell rings must sign in at the office and receive an admit-to-class slip.""",
                    normal,
                ),
            ]
        )
    )

    Story.append(
        KeepTogether(
            [
                Paragraph("Teacher Conferences", h2),
                Paragraph(
                    """Conferences to discuss each student's progress are held every year in November. Sign up usually after Back-to-School night through email and the Signup Genius program. Teachers may also send home a letter suggesting a time and day. Conferences are with the teacher, but the principal can be present if requested. Additionally, the school emphasizes that you may arrange a meeting with the teacher or principal at any time about anything that concerns your students wellbeing and education.""",
                    normal,
                ),
            ]
        )
    )

    Story.append(
        KeepTogether(
            [
                Paragraph("Telephones", h2),
                Paragraph(
                    """Each classroom has a telephone, and a telephone for essential calls is available for students in the lobby, on the first floor. Students may use the office telephone in an emergency. Social arrangements should be made at home.""",
                    normal,
                ),
            ]
        )
    )

    if False:
        Story.append(
            KeepTogether(
                [
                    Paragraph("Testing", h2),
                    Paragraph(
                        """Students in 2nd grade are screened for "giftedness". In addition, Maryland tests its students using the PARCC (Partnership for Assessment for Readiness for College and Careers). Testing protocols and frequency are being changed, please visit Testing Information in the Parents section at the MCPS website for the most recent information.""",
                        normal,
                    ),
                ]
            )
        )

    Story.append(
        KeepTogether(
            [
                Paragraph("Transfers and Withdrawals", h2),
                Paragraph(
                    """In certain circumstances, parents may request the transfer of a student from one school to another. Forms are available in the office. The school should be notified promptly if a student must be withdrawn or transferred to another school. This policy applies for withdrawal during a school year, as well as at the end of the school year.""",
                    normal,
                ),
            ]
        )
    )

    # "Today is Newsday (The TIN) - PTA e-news letter"
    # The PTA publishes a weekly news bulletin called Today is Newsday (TIN). Distributed by email via the listserv on sundays, the TIN contains important a ouncements and user reminders of upcoming events. News and information for publication may be submitted by email to: tin@somersetpta.org. Sign up to receive the TIN electronically sendingUbscribe@yahoogroups.com. your email address to somerset-net-

    # Story.append(
    #     KeepTogether(
    #         [
    #             Paragraph("Used Book Sale and Bake Sale", h2),
    #             Paragraph(
    #                 """Somerset students colar rods from their homes and community to be she Spring. Prin prices during this annual tvo-day event in the summer izes are given to the classes Collecting the largest nither of books. The Bake Sale, which is held in conjunction with the Used Book Sale, features a wide variety of homemade foods to sustain the book buyers. browsers and sellers.""",
    #                 normal,
    #             ),
    #         ]
    #     )
    # )

    Story.append(
        KeepTogether(
            [
                Paragraph("Valentine's Day", h2),
                Paragraph(
                    """On February 14 (or the school day before if this date falls on weekend), students may bring Valentine cards to exchange with all of their classmates, and a class celebration may follow. Contact your room parents for details.""",
                    normal,
                ),
            ]
        )
    )

    Story.append(
        KeepTogether(
            [
                Paragraph("Volunteering", h2),
                Paragraph(
                    """To volunteer at the school or in your classroom please, contact your teacher or specials teachers. There are many PTA events throughout the year that can use your help from the Back to School Classic Race to our book fairs and other community events. The PTA also has opportunities for parents to help at recess and/or lunch. Volunteers will need to complete the online MCPS Child Abuse and Neglect recognition training found the MCPS website http://www.montgomeryschoolsmd.org/childabuseandneglect/""",
                    normal,
                ),
                url2qr("http://www.montgomeryschoolsmd.org/childabuseandneglect/"),
                Paragraph(
                    """Volunteers who will be attending extended day field trips wil need to complete a finger printing and background check. Please ask your teacher or the principal's office about these requirements. You can also read more on these policies on the Montgomery County Public School FAQ at: http://www.montgomeryschoolsmd.org/uploadedFiles/childabuseandneglect/160902-ChildAbuseVolunteer-FAQs.pdf""",
                    normal,
                ),
                url2qr(
                    "http://www.montgomeryschoolsmd.org/uploadedFiles/childabuseandneglect/160902-ChildAbuseVolunteer-FAQs.pdf"
                ),
                #                Paragraph(
                #                    """Volunteers will need to complete the online MCPS Child Abuse and Neglect recognition training found the MCPS websitehttp://www.montgomeryschoolsmd.org/childabuseandneglect/""",
                #                    normal,
                #                ),
                #                Paragraph(
                #                    """Volunteers who will be attending extended day field trips wil need to complete a finger printing and background check. Please ask your teacher or the principal's office about these requirements. You can also read more on these policies on the Montgomery County Public School FAQ at:http://www.montgomeryschoolsmd.org/uploadedFiles/childabuseandneglect/160902-ChildAbuse-Volunteer-FAQs.pdf""",
                #                    normal,
                #                ),
            ]
        )
    )

    Story.append(
        KeepTogether(
            [
                Paragraph("Weapons and Pocketknives", h2),
                Paragraph(
                    """Students must not bring anything to school that may cause injury, or can be construed as a weapon, such as Swiss Army knives or small pocketknives, toy weapons or dangerous liquids. (Disciplinary action may be taken, including Suspension.)""",
                    normal,
                ),
            ]
        )
    )

    Story.append(
        KeepTogether(
            [
                Paragraph("Weather Contingency Plan", h2),
                Paragraph(
                    """If school is closed for more than five days during the school year due to weather emergencies, the Weather Contingency Plan may be implemented and additional student instructional days may be added to the school year. Visit the MCPS website for schedule changes due to weather.""",
                    normal,
                ),
            ]
        )
    )

    if False:
        Story.append(
            KeepTogether(
                [
                    Paragraph("Websites for Somerset and the PTA", h2),
                    Paragraph(
                        """The PTA website is https://somersetelementary.memberhub.com/""",
                        normal,
                    ),
                    url2qr("https://somersetelementary.memberhub.com/"),
                    Paragraph(
                        """The Somerset Elementary MCPS website is https://www.montgomeryschoolsmd.org/schools/somersetes Links include the Media Center, Counseling, Specialists and Classrooms that are updated throughout the year. The Staff Directory link takes you to Somerset's online telephone and email directory. The MCPS Home link at the bottom of the page takes you to the Montgomery County Public School website for comprehensive information.""",
                        normal,
                    ),
                    url2qr("https://www.montgomeryschoolsmd.org/schools/somersetes"),
                    #                Paragraph(
                    #                    """Links include the Media Center, Counseling, Specialists and Classrooms that are updated throughout the year. The Staff Directory link takes you to Somerset's online telephone and email directory. The MCPS Home link at the bottom of the page takes you to the Montgomery County Public School website for comprehensive information.""",
                    #                    normal,
                    #                ),
                ]
            )
        )

    Story.append(
        KeepTogether(
            [
                Paragraph("Yearbook", h2),
                Paragraph(
                    """Somersets yearbook is published and available for purchase near the end of each year, with photos of all students and staff along with pictures of major school events.""",
                    normal,
                ),
            ]
        )
    )

    Story.append(PageBreak())

    linkedHeading(Story, "Q & A", toch1)

    # Somerset FAQ — questions asked and answered!
    Story.append(Paragraph("""Q: What should I do when my child is sick?""", h2))
    Story.append(
        Paragraph(
            """A: Somerset loves seeing your kids, but please keep them home when they are sick: they must be fever-free for 24 hours to return. For strep throat and other infections requiring antibiotics, please check with your healthcare provider about when it is safe to return to school. If a communicable disease has been diagnosed or lice nits have been found, please notify our health room by calling 240-740-1102.""",
            normal,
        )
    )

    Story.append(Paragraph("""Q: When does school start? When does school end?""", h2))
    Story.append(
        Paragraph(
            """A: Before 8:40 am, there is no supervision available for children, and children will not be permitted to enter the building, even during inclement weather. THE ONLY EXCEPTIONS ARE children enrolled in Bar-T before care or children who are registered for a morning club with Enrichment Academies. Students are dismissed starting at 3:22. Walkers and car riders go first, then bus riders. Kindergarteners not riding the bus must be picked up at their classroom door in the rear.""",
            normal,
        )
    )

    Story.append(
        KeepTogether(
            [
                Paragraph(
                    """Q: What if I need childcare before or after school?""", h2
                ),
                Paragraph(
                    """A: Bar-T provides before and/or aftercare for a fee. Bar-T Kids Club at 240-364-4196 https://www.bar-t.com/program/kids-club/""",
                    normal,
                ),
                url2qr("https://www.bar-t.com/program/kids-club/"),
            ]
        )
    )

    Story.append(Paragraph("""Q: What should I do when my child is late?""", h2))
    Story.append(
        Paragraph(
            """A: The first bell is at 8:55. The second bell is at 9:00 am. By 9:00, students are expected to be in their classrooms. If your child arrives at school at 9:00 or after, please accompany him or her into the main office and sign them in on the sign in sheet. Students need a tardy slip to go to class.""",
            normal,
        )
    )

    Story.append(
        Paragraph(
            """Q: What should I do if my child has a medical appointment during the school day?""",
            h2,
        )
    )
    Story.append(
        Paragraph(
            """A: When possible, please schedule medical, dental, and orthodontist appointments outside of school hours or if that is not possible, consider scheduling during lunch and recess to minimize the loss of school time. Please email or send a note to your child's teacher letting them know. You will need to come into school to sign out your child and also to sign your child back in.""",
            normal,
        )
    )

    Story.append(
        Paragraph(
            """Q: What should I do if my child forgets his/her lunch or homework or musical instrument, etc.?""",
            h2,
        )
    )
    Story.append(
        Paragraph(
            """A: The main office has a bin for forgotten items. You may email your child's teacher so that your child will know that the item is in the main office. The main office does not call children out of their classrooms to pick up missing items.""",
            normal,
        )
    )

    Story.append(
        Paragraph(
            """Q: What should I do if my child is going to go home with a friend?""",
            h2,
        )
    )
    Story.append(
        Paragraph(
            """A: Please send a note with your child and/or email the child's teacher before 12 noon. If your child is riding the bus, please send a note for the bus driver as well.""",
            normal,
        )
    )

    # """Q: How do I know if school is closed or delayed when the weather is bad?"""
    # """A: Mrs. Morris will send out a message over the email system and somerset-net listserv. There is also a county email system that you can access at www.montgomeryschoolsmd.org . Subscribe to MCPS QuickNotes for weather-related email messages. closing information for Bar-T Kids Club call 240-364 4196"""

    Story.append(Paragraph("""Q: What if my child gets sick in school?""", h2))
    Story.append(
        Paragraph(
            """A: Your child will be sent to the health room. In the event of tever or vomiting, you will be called. Please make sure it your child has allergies or asia and requires any emergency medications that you have a medication administration form completed by your doctor and the appropriate medicatom stored in the health room.""",
            normal,
        )
    )

    Story.append(
        KeepTogether(
            [
                Paragraph("""Q: What if my child is being bullied?""", h2),
                Paragraph(
                    """A: Please contact Principal Wiebe, the principal, or Ms. Musser, the school counselor, to discuss any bullying situation. Most can be resolved with simple intervention. If it is happening at recess, the paraeducators who monitor recess can be asked to assist. To learn more about reporting bullying, harassment or intimidation and see a copy of the reporting form please visit the MCPS web site: http://www.montgomeryschoolsmd.org/departments/forms/pdf/230-35.pdf""",
                    normal,
                ),
                url2qr(
                    "http://www.montgomeryschoolsmd.org/departments/forms/pdf/230-35.pdf"
                ),
            ]
        )
    )
    #    Story.append(
    #        Paragraph(
    #            """To learn more about reporting bullying, harassment or intimidation and see a copy of the reporting form please visit the MCPS web site: http://www.montgomeryschoolsmd.org/departments/forms/pdf/230-35.pdf""",
    #            normal,
    #        )
    #    )

    Story.append(Paragraph("""Q: What is the policy for recess?""", h2))
    Story.append(
        Paragraph(
            """A: Recess is held every day, except for half days. In weather above 32 degrees, the children generally play outside. On inclement weather days, recess is held indoors in the classroom. Due to supervision constraints, two classes are usually combined for recess. The PTA has provided board games for indoor recess and equipment for outdoor recess.""",
            normal,
        )
    )

    Story.append(Paragraph("""Q: What is the policy for snacks?""", h2))
    Story.append(
        Paragraph(
            """A: Students in all grades have a designated snack time. Please send foods that a healthy and have a minimum of noise and mess. Please remember that we are a nut free school — no peanut butter or peanut products or tree nuts.""",
            normal,
        )
    )

    Story.append(
        KeepTogether(
            [
                Paragraph("""Q: Does Somerset offer after school activities?""", h2),
                Paragraph(
                    """A: Yes. We have a wide variety of before and after school programs offered through Enrichment Academies. Clubs are offered for three "semesters" each year, fall, winter, and spring. There is a registration period. Please visit https://somerset.enrichment-academies.com/ to learn more. Scholarships are offered based on specific need.""",
                    normal,
                ),
                url2qr("https://somerset.enrichment-academies.com/"),
            ]
        )
    )

    Story.append(Paragraph("""Q: How does discipline work at Somerset?""", h2))
    Story.append(
        Paragraph(
            """A: Each teacher has his or her own classroom method for handling disruptive behavior, involving warnings and consequences, as well as opportunities to earn preferred activity points and other perks for good behavior. The lunchroom uses a table points system. Students can earn good behavior rewards, such as crazy hair and crazy sock days, for accumulating school-wide good behavior. Our disciplinary actions follow the MCPS student code of conduct. Copies of Somerset's discipline policy are available in the school office or the Somerset Elementary website.""",
            normal,
        )
    )

    Story.append(
        Paragraph("""Q: How will I know what my child is working on in class?""", h2)
    )
    Story.append(
        Paragraph(
            """A: Most grades and also individual teachers send out periodic emails giving parents an overview of the projects being done and the subject matter covered. Work, such as graded homework, exit cards, papers, and tests are also returned in each students' folder. If you have questions, please email your child's teacher. Teachers respond to email within 24-48 hours.""",
            normal,
        )
    )

    Story.append(
        Paragraph("""Q: How are drop off and dismissal run and supervised?""", h2)
    )
    Story.append(
        Paragraph(
            """A: Students line up outside of their classrooms beginning at 8:40; students are not permitted on the field at this time. In the afternoon, students not riding the bus are dismissed in three separate groups, After School Care (Bar-T), walkers and car riders, and after school clubs. Each bus has at least one fifth grade safety patrol rider who supervises the bus. Buses are usually met by the principal and assistant principal or other staff members in the morning. Bus riders are dismissed individually by bus in the afternoon. Kindergarten riders are dismissed first and board first. Parents or caregivers are expected to meet the bus in the afternoon, particularly for grades K-2.""",
            normal,
        )
    )

    Story.append(
        Paragraph("""Q: What are the opportunities to volunteer in school?""", h2)
    )
    Story.append(
        Paragraph(
            """A: Somerset is happy to have the help und support of parents. Some classes have volunteer opportunities. Other teachers need help with specific projects or even organizing papers and supplies. Check with your teacher about needs in your classroom.""",
            normal,
        )
    )

    Story.append(
        Paragraph(
            """In addition, there are many PTA events throughout the year that can use your help. Volunteers in the classroom will need to complete the online MCPS Child Abuse and Neglect recognition training found http://www.montgomeryschoolsmd.org/childabuseandneglect/ Volunteers who will be attending extended day field trips will need to complete a finger printing and background check. Please ask your teacher or the principal's office about these requirements for your volunteering. You can also read the MCPS FAQ at: http://www.montgomeryschoolsmd.org/uploadedFiles/childabuseandneglect/160902-ChildAbuse-Volunteer-FAQs.pdf """,
            normal,
        )
    )
    #  These events are posted on the Website and in the TIN. www.somersetpta.com

    Story.append(
        Paragraph(
            """Q: What is the difference between the PTA and the Foundation?""", h2
        )
    )
    Story.append(
        Paragraph(
            """A: The PTA provides direct support for teachers and students in the classroom, in the form of annual teacher stipends, scholarships and other financial support for students in need, and basic supplies, as well as providing grants for specific activities, such as assemblies or teacher professional development. It provides a cultural arts program for students and is responsible for maintaining a robust afterschool enrichment activities program. It also organizes and hosts all major community events at the school throughout the year, from the 8k Race to International Night. It serves as an advocate for families, students, and teachers within the school and within the entire BCC cluster. It provides multiple forums for communications with families and the school and hosts regular meetings. The PTA Board of Directors includes parents elected by the PTA, the principal, and a teacher representative.""",
            normal,
        )
    )
    Story.append(
        Paragraph(
            """The Somerset Foundation focuses on large-scale capital and technological projects to improve academic and physical features at the school. The Board consists of nominated parents and community leaders, along with the principal, PTA President, and a teacher representative. The Board raises funds from parents and the community for a variety of projects. Early projects included the underwriting of the arts initiative; the creation and enhancement of the original computer lab; the development of the service learning curriculum; and the purchase of classroom books, as well as enhancements to the school when it was renovated. Most recently, the Foundation was focused on the installation of the new turf field with the Field Committee.""",
            normal,
        )
    )

    Story.append(PageBreak())

    if True:
        Story.append(PageBreak())

        linkedHeading(Story, "Staff Directory", toch1)

        url1 = "https://www.montgomeryschoolsmd.org/schools/somersetes/staff/directory/"
        link1 = f"<link href='{url1}'>{url1}</link>"
        Story.append(Paragraph(link1, centered_style))
        Story.append(url2qr(link1))

        for staff_member in staff_order:
            staff_table = []
            staff_name = staff_member.get("formal")
            staff_nickname = staff_member.get("nickname")
            staff_title = staff_member.get("title")
            staff_email = staff_member.get("email")
            if not staff_email:
                print(f"WARN#2 no email known for {staff_member=}")

            name_row = [
                Paragraph(staff_name),
            ]
            if staff_nickname:
                name_row.append(Paragraph(f"({staff_nickname})", style_right))
            staff_table.append(name_row)

            staff_table.append(
                [
                    Paragraph(staff_title),
                    Paragraph(format_email(staff_email), style_right),
                ]
            )
            t = Table(
                staff_table,
                style=[
                    ("LINEBELOW", (-2, -1), (-1, -1), 0.25, colors.black),
                    # ("GRID", (0, 0), (-1, -1), 0.5, colors.black),
                ],
            )

            Story.append(KeepTogether(t))

    Story.append(PageBreak())

    psr = pool_to_student_relations(pool)
    num_students = 0

    by_lastname = {}
    by_firstname = {}
    by_street = {}
    by_homeroom = {}

    for student_uid, student in psr.items():
        num_students += 1
        student_name = student["Student"]

        lastname, firstname = student_name.split(", ")
        by_lastname.setdefault(lastname, []).append(student_uid)
        by_firstname.setdefault(firstname, []).append(student_uid)

        address1 = student.get("Address1")
        street_name = get_street(address1)
        if street_name:
            by_street.setdefault(street_name, []).append(student_uid)
            # I would prefer perfect sorting of addresses, but too many records have 1 child withheld, while the other is given an address

        grade = get_grade(student)
        teacher = get_teacher(student)
        homeroom_key = f"{grade} {teacher}"
        by_homeroom.setdefault(homeroom_key, []).append(student_uid)

    linkedHeading(Story, "Full Details by Last Name", toch1)

    for student_uid in psr:
        student = psr[student_uid]
        num_students += 1
        student_name = student["Student"]

        kt = []

        lastname, firstname = student_name.split(", ")
        student_anchor = f"<a name='{student_uid}'/>{student_name}"
        kt.append(Paragraph(student_anchor, details_student_name_style))

        grade = get_grade(student)
        teacher = get_teacher(student)
        aclass_uid = class_uid(grade=grade, teacher=teacher, entry=student)
        class_anchor = (
            f"<link href='#{aclass_uid}'>Grade: {grade} Teacher: {teacher}</link>"
        )
        kt.append(Paragraph(class_anchor, details_class_teacher_style))

        if phone := student.get("Phone"):
            kt.append(
                Paragraph(
                    f"Phone: {format_phone_link(format_phone(phone))}",
                    details_phone_style,
                )
            )
        address1 = student.get("Address1")
        address2 = student.get("Address2")
        if address1 or address2:
            # address = f"{student.get('Address1','')}<br/>{student.get('Address2','')}"
            address = format_address(student)
            kt.append(Paragraph(address, details_address_style))

        street_name = get_street(address1)

        data = []

        data_keys = []
        for relation in student["Relations"]:
            for key, value in relation.items():
                if value != withheld_marker:
                    if key not in data_keys:
                        data_keys.append(key)

        offset_width = 10
        # relationship_width = 0.65
        for relation in student["Relations"]:
            data_row = []
            any_values = [x for x in relation.values() if x != withheld_marker]
            if any_values:
                for key in data_keys:
                    value = relation.get(key)
                    if value and value != withheld_marker:
                        data_row.append(
                            Paragraph(value, styleSheet["BodyText"]),
                        )
                    else:
                        data_row.append(None)
            if data_row:
                ready = [None]
                ready.extend(data_row)
                data.append(ready)
        if data:
            num_cols = len(data[0])
            # col_width = (5 * inch - offset_width) / (num_cols - 1)
            # col_widths = [.75*inch] + [col_width] * (num_cols-1)
            col_widths = [offset_width] + [None] * (num_cols - 2)
            t = Table(
                data,
                hAlign="RIGHT",
                # colWidths=[2.4 * inch, 2.5 * inch, 2.5 * inch],
                colWidths=col_widths,
                style=[
                    ("LINEABOVE", (1, 0), (-1, -1), 0.25, colors.black),
                    ("LINEBELOW", (1, 0), (-1, -1), 0.25, colors.black),
                    ("LINEBEFORE", (2, 0), (-1, -1), 0.25, colors.black),
                    # ("GRID", (0,0), (-1, -1), 0.5, colors.black),
                    ("VALIGN", (0, 0), (-1, -1), "TOP"),
                    ("LEFTPADDING", (0, 0), (0, -1), 0),
                    ("LEFTPADDING", (1, 0), (-1, -1), 1),
                    # ('LEFTPADDING', (-1,0), (-1, -1), 1),
                    ("RIGHTPADDING", (0, 0), (-1, -1), 1),
                    ("TOPPADDING", (0, 0), (-1, -1), 0),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 0),
                    # ('FONTSIZE', (0, 0), (-1, -1), 16),
                    # ('leftIndent', (0, 0), (1, -1), 10)
                ],
            )

            kt.append(t)
        Story.append(KeepTogether(kt))

    Story.append(PageBreak())

    ptext = "By Grade & Teacher"
    linkedHeading(Story, ptext, toch1)
    Story.append(Spacer(1, 12))

    tgs = pool_to_teacher_grade_student_uids(pool)
    for grade in tgs:
        for teacher in tgs[grade]:
            aclass_uid = class_uid(grade=grade, teacher=teacher)
            class_text = f"<a name='{aclass_uid}'/>{grade} {teacher}"
            agroup = []
            agroup.append(Paragraph(class_text, teacher_style))

            teacher_email = get_teacher_email(teacher)
            if teacher_email:
                agroup.append(Paragraph(teacher_email, teacher_email_style))
            else:
                print(f"WARN#1 no email known for {grade=} {teacher=}")

            for student_uid in tgs[grade][teacher]:
                student = psr[student_uid]
                student_link = f"\u2022 <link href='#{student_uid}'>{student.get('Student')}</link>"
                p = Paragraph(student_link, student_teacher_style)
                agroup.append(p)

            Story.append(KeepTogether(agroup))

    Story.append(PageBreak())

    ptext = "By First Name"
    linkedHeading(Story, ptext, toch1)

    name_flow = []
    for firstname in sorted(by_firstname):
        for student_uid in by_firstname[firstname]:
            student = psr[student_uid]
            student_name = student.get("Student")
            alastname, afirstname = student_name.split(", ")

            student_link = (
                f"<link href='#{student_uid}'>{afirstname} {alastname}</link>"
            )
            p = Paragraph(student_link, styleSheet["BodyText"])
            name_flow.append(p)
    Story.append(BalancedColumns(name_flow))

    Story.append(PageBreak())

    ptext = "By Street"
    linkedHeading(Story, ptext, toch1)
    Story.append(Spacer(1, 12))

    for street_name in sorted(by_street):

        astreet_url = street_url(street_name)
        if astreet_url:
            street_anchor = f"<a name='{astreet_url}'/>{street_name}"
        else:
            street_anchor = street_name

        Story.append(Paragraph(street_anchor, h2))
        for student_uid in by_street[street_name]:
            student = psr[student_uid]
            student_name = student.get("Student")
            student_link = f"\u2022 <link href='#{student_uid}'>{student_name}</link>"
            p = Paragraph(student_link, student_street_style)
            Story.append(p)

    Story.append(PageBreak())

    ptext = "About This Directory"
    linkedHeading(Story, ptext, toch1)
    Story.append(Spacer(1, 12))

    url1 = "https://www.montgomeryschoolsmd.org/departments/forms/detail.aspx?formID=475&formNumber=281-13"
    link1 = f"<link href='{url1}'>{url1}</link>"

    cover_credit_url = "https://www.instagram.com/maryvinogradfineart/"

    Story.append(
        KeepTogether(
            [
                Paragraph("""Cover Artwork by Mary Vinograd"""),
                Paragraph(cover_credit_url, centered_style),
                Spacer(1, 12),
                url2qr(cover_credit_url),
                Spacer(1, 12),
                Spacer(1, 12),
                Paragraph(
                    """
                The information in this directory is derived from information collected by MCPS. Exclusion or corrections can be made via Form number: 281-13 and is available online at
                """
                ),
                Spacer(1, 12),
                Paragraph(link1, centered_style),
                Spacer(1, 12),
                url2qr(link1),
            ]
        )
    )

    return Story


def story_to_pdf(Story, owner=None, filename="mypdf1.pdf"):
    tmppdf = tempfile.NamedTemporaryFile(suffix=".pdf")
    doc = MyDocTemplate(tmppdf.name)
    if owner:
        doc.owner = owner

    success = False
    try:
        doc.multiBuild(Story)
        success = True
    except reportlab.platypus.doctemplate.LayoutError as e:
        print(e)
        breakpoint()
        # pdb.set_trace()
        while not success:
            try:
                doc.build(Story)
            except Exception as e:
                print("Unexpected error:", sys.exc_info()[0])
                print("exception:", e)
                print("error:", sys.exc_info()[1])
                traceback.print_exc(file=sys.stdout)
                print("removing %s" % Story[0])
                Story = Story[1:]

    if success:
        from shutil import copyfile

        copyfile(tmppdf.name, filename)
    else:
        print(f"failed to make {filename}")

    return


withheld_marker = "(withheld)"


def xlsx_to_pool(src, sheet=None):
    if sheet is None:
        from openpyxl import load_workbook

        wb = load_workbook(filename=src)
        sheet = wb.active
    else:
        pass

    # col_labels = sheet.row_values(1)
    col_labels = [x.value for x in next(sheet.rows)]

    col_clean_labels = [x.strip() for x in col_labels]
    clean_col = {x: y for x, y in zip(col_labels, col_clean_labels)}

    # num_cols = len(col_labels)

    Directory_Withholding_key = "Directory Withholding-YN"
    if Directory_Withholding_key not in col_labels:
        print(f"{Directory_Withholding_key} was not found. not safe to load this")
        return

    # nuke the preK

    num_withheld = 0
    num_accepted = 0
    pool = []

    emails_with_includes = {}
    emails_with_excludes = {}

    for araw in sheet.rows:

        raw_dict = {}
        for label, acell in zip(clean_col, araw):
            raw_dict[label] = str(acell.value)

        adict = {clean_col[x]: str(y) for x, y in raw_dict.items()}
        withheld = False

        delkeys = []
        for k in adict.keys():
            if adict[k] is None:
                delkeys.append(k)
        for k in delkeys:
            del adict[k]
        Directory_Withholding = adict.get(Directory_Withholding_key)

        if Directory_Withholding != "N":
            withheld = True
            # for k in [
            #     "Sch Num",
            #     "School",
            #     "Birth Date",
            #     #'Directory Withholding-YN',
            #     "Phone",
            #     "Address1",
            #     "Address2",
            #     "Relation",
            #     "Name",
            #     "Cell Phone",
            #     "Email",
            #     'Student',
            #         #'Homeroom Teacher', 'Grade',
            # ]:
            #     #adict[k] = withheld_marker
            #     continue
            if Directory_Withholding == "Y":
                pass
            elif "withholding" in Directory_Withholding.lower():
                # withheld = True
                print(f"skipping header based on seeing '{Directory_Withholding}'")
            else:
                print(
                    f"{Directory_Withholding_key} = '{Directory_Withholding}' ... not understood, so dropping this record"
                )

        Grade = get_grade(adict)
        if Grade == "SE PreK":
            # special case exclusion
            continue

        if not withheld:
            pool.append(adict)
            num_accepted += 1
        else:
            num_withheld += 1

        if email := adict.get("Email"):
            if withheld:
                emails_with_excludes[email] = True
            else:
                emails_with_includes[email] = True

    print(f"{num_withheld=} {num_accepted=}")

    for email in sorted(emails_with_excludes):
        if emails_with_includes.get(email):
            print(f" ... FYI email {email} had includes and excludes")

    # reminder the 'lower' is necessary for names like "de Bruin" with an initial lowercase
    ordered_pool = [
        k
        for k in sorted(
            pool, reverse=False, key=lambda item: item.get("Student").lower()
        )
    ]

    return ordered_pool


def filter_pool_to_students(pool, students):

    out_pool = []
    for arec in pool:
        if arec.get("Student") in students:
            out_pool.append(arec)
    return out_pool


def xlsx_to_emails(src):
    """does not respect witholding, does not need to as currently used"""
    from openpyxl import load_workbook

    wb = load_workbook(filename=src)
    sheet = wb.active

    col_labels = []
    for row in sheet.iter_rows(min_row=0, min_col=0, max_row=1, max_col=4000):
        for cell in row:
            # print(cell.row, cell.column, cell.value)
            val = cell.value
            if val:
                val = val.strip()
            col_labels.append(val)
    while col_labels[-1] is None:
        col_labels.pop()
    num_cols = len(col_labels)

    # num_withheld = 0
    # num_accepted = 0

    preapproved = set(
        [
            "Cariaso, Hana",
            "Levitas, Spencer Nathan",
            "Press, Hari Richard Singh",
            "Lawler, Amelia Marie",
            "Lawler, Celine Kimbell",
        ]
    )
    emails = {}
    for row in sheet.iter_rows(
        min_row=2,
        min_col=0,
        # max_row=6,
        max_col=num_cols,
    ):
        adict = dict(zip(col_labels, [x.value for x in row]))
        email = get_relation_email(adict)
        if email:
            email = email.lower()
            student = adict.get("Student")
            emails.setdefault(email, copy(preapproved)).add(student)

    return emails


def format_phone(phone):
    if phone:
        if m := re.search(r"^(\d{3})(\d{3})(\d{4})$", phone):
            out = m.group(1) + "-" + m.group(2) + "-" + m.group(3)
            return out

    return phone


def street_url(street_name):
    if street_name:
        return "street_" + hashlib.sha1(street_name.encode("utf-8")).hexdigest()
    else:
        return None


def format_address(student):

    address1 = student.get("Address1")
    address2 = student.get("Address2")

    street_name = get_street(address1)
    url = street_url(street_name)

    link = f"<link href='#{url}'>{street_name}</link>"
    pretty_address1 = re.sub(street_name, link, address1)

    if address1 and address2:
        address = f"{pretty_address1}<br/>{address2}"
    else:
        address = f"{student.get('Address1','')}<br/>{student.get('Address2','')}"
    return address


def format_email(email):
    return f'<a href="mailto:{email}">{email}</a>'


def format_phone_link(phone):
    return f'<a href="tel:{phone}">{phone}</a>'


def get_teacher_email(teacher):
    for staff_member in staff_order:
        if staff_member.get("directory_key") == teacher:
            return staff_member.get("email")
    return None


staff_order = [
    {
        "formal": "Mr. Travis J Wiebe",
        "title": "Principal",
        "email": "Travis_J_Wiebe@mcpsmd.org",
    },
    {
        "formal": "Mrs. Bess W Treat",
        "title": "Assistant School Administrator",
        "email": "Bess_W_Treat@mcpsmd.org",
    },
    {
        "formal": "Mrs. Nancy L Conway",
        "title": "School Secretary",
        "email": "Nancy_L_Conway@mcpsmd.org",
    },
    {
        "formal": "Ms. Susan E Stringham",
        "title": "School Admin Secretary",
        "email": "Susan_Stringham@mcpsmd.org",
    },
    {
        "formal": "Mrs. Beth G Andreassi",
        "title": "Paraeducator",
        "email": "Beth_G_Andreassi@mcpsmd.org",
    },
    {
        "formal": "Ms. Megan C Appleton",
        "directory_key": "Appleton, Megan",
        "nickname": "Meg",
        "title": "Teacher, Grade 2",
        "email": "Megan_C_Appleton@mcpsmd.org",
    },
    {
        "formal": "Ms. Ehlam Aslam",
        "directory_key": "Aslam, Ehlam",
        "title": "Teacher, Grade 1",
        "email": "Ehlam_Aslam@mcpsmd.org",
    },
    {
        "formal": "Mrs. Elissa M Bean",
        "title": "Teacher, ELD",
        "email": "Elissa_M_Bean@mcpsmd.org",
    },
    {
        "formal": "Ms. Barbara A Berlin",
        "directory_key": "Berlin, Barbara",
        "title": "Teacher, Grade 4",
        "email": "Barbara_A_Berlin@mcpsmd.org",
    },
    {
        "formal": "Ms. Linda J Bryant",
        "title": "Teacher, General Music",
        "email": "Linda_J_Bryant@mcpsmd.org",
    },
    {
        "formal": "Mrs. HeeJung Burns",
        "title": "Teacher, ELD",
        "email": "HeeJung_Burns@mcpsmd.org",
    },
    {
        "formal": "Ms. Merritt M Crowder",
        "title": "Media Specialist",
        "email": "Merritt_M_Crowder@mcpsmd.org",
    },
    {
        "formal": "Ms. Marynell A Curtis",
        "title": "Teacher, Art",
        "email": "Marynell_A_Curtis@mcpsmd.org",
    },
    {
        "formal": "Mrs. Antoinette D Davidov",
        "directory_key": "Davidov, Antoinette",
        "nickname": "Annie",
        "title": "Teacher, Kindergarten",
        "email": "Antoinette_D_Davidov@mcpsmd.org",
    },
    {
        "formal": "Barbara Davis",
        "title": "Teacher, Resource",
        "email": "Barbara_C_Davis@mcpsmd.org",
    },
    {
        "formal": "Mrs. Danielle B Ellis",
        "title": "Reading Specialist",
        "email": "Danielle_B_Ellis@mcpsmd.org",
    },
    {
        "formal": "Mr. Todd G Ellis Jr",
        "nickname": "TJ",
        "title": "Teacher, Physical Education",
        "email": "Todd_G_EllisJr@mcpsmd.org",
    },
    {
        "formal": "Mrs. Anne E Flores",
        "nickname": "Brooke",
        "title": "Teacher, Staff Development",
        "email": "Anne_E_Flores@mcpsmd.org",
    },
    {
        "formal": "Miss Emily Freilich",
        "directory_key": "Freilich, Emily",
        "title": "Speech Pathologist",
        "email": "Emily_Freilich@mcpsmd.org",
    },
    {
        "formal": "Ladraine Greene",
        "title": "Teacher, Resource",
        "email": "Ladraine_L_Greene@mcpsmd.org",
    },
    {
        "formal": "Ms. Karen L Hansel",
        "directory_key": "Hansel, Karen",
        "title": "Teacher, Grade 1",
        "email": "Karen_L_Hansel@mcpsmd.org",
    },
    {
        "formal": "Mr. Michael D Johnson",
        "title": "Building Services Worker",
        "email": "Michael_D_Johnson@mcpsmd.org",
    },
    {
        "formal": "Ms. Shana M Joyce",
        "directory_key": "Joyce, Shana",
        "title": "Teacher, Kindergarten",
        "email": "Shana_M_Joyce@mcpsmd.org",
    },
    {
        "formal": "Mrs. Amanda M Kim",
        "title": "Teacher, Instrumental Music",
        "email": "Amanda_Kim@mcpsmd.org",
    },
    {
        "formal": "Mr. Gregory P Matwey",
        "directory_key": "Matwey, Gregory",
        "nickname": "Greg",
        "title": "Teacher, Grade 5",
        "email": "Gregory_P_Matwey@mcpsmd.org",
    },
    {
        "formal": "Ms. Tiffany A Mclean",
        "title": "Media Assistant",
        "email": "Tiffany_A_Mclean@mcpsmd.org",
    },
    {
        "formal": "Mrs. Katherine G Musser",
        "nickname": "Kate",
        "title": "Counselor",
        "email": "Katherine_G_Musser@mcpsmd.org",
    },
    {
        "formal": "Mr. Daniel J Oddo",
        "nickname": "Dan",
        "title": "Teacher, Resource",
        "email": "Daniel_Oddo@mcpsmd.org",
    },
    {
        "formal": "Dr. Tiffany E Proctor",
        "directory_key": "Proctor, Tiffany",
        "title": "Teacher, Grade 4",
        "email": "Tiffany_E_Proctor@mcpsmd.org",
    },
    {
        "formal": "Mrs. Meghan M Rivera",
        "directory_key": "Rivera, Meghan",
        "title": "Teacher, Grade 1",
        "email": "Meghan_M_Rivera@mcpsmd.org",
    },
    {
        "formal": "Mrs. Regina M Sakaria",
        "directory_key": "Sakaria, Regina",
        "title": "Teacher, Grade 3",
        "email": "Regina_Sakaria@mcpsmd.org",
    },
    {
        "formal": "Ms. Mary Agnes S Sisti",
        "directory_key": "Sisti, Mary Agnes",
        "nickname": "Maggie",
        "title": "Teacher, Grade 3",
        "email": "MaryAgnes_S_Sisti@mcpsmd.org",
    },
    {
        "formal": "Mrs. Diane M Smith",
        "title": "Teacher, Grade 5",
        "directory_key": "Smith, Diane",
        "email": "Diane_M_Smith@mcpsmd.org",
    },
    {
        "formal": "Mr. William A Thompson Jr",
        "directory_key": "Thompson Jr, William",
        "nickname": "Billy",
        "title": "Teacher, Grade 5",
        "email": "William_A_ThompsonJr@mcpsmd.org",
    },
    {
        "formal": "Mrs. Kathryn L Truppner",
        "title": "Paraeducator",
        "email": "Kathryn_L_Truppner@mcpsmd.org",
    },
    {
        "formal": "Ms. Dana Ward",
        "directory_key": "Ward, Dana",
        "title": "Teacher, Grade 2",
        "email": "Dana_Ward@mcpsmd.org",
    },
    # {
    #    "formal": "Kate Waldmann",
    #    "title": "Special Education Paraeducator",
    #    # "email": "Michael_D_Johnson@mcpsmd.org",
    # },
    {
        "formal": "Mrs. Elahe Yazdantalab",
        "title": "Paraeducator",
        "email": "Elahe_Yazdantalab@mcpsmd.org",
    },
    {
        "formal": "Ms. Danielle C McIntyre-Still",
        "title": "Building Services Manager",
        "email": "Danielle_C_McIntyre-Still@mcpsmd.org",
    },
    {
        "formal": "Mrs. Maria M Portillo-Coreas",
        "title": "Building Service Worker Sh 1",
        "email": "Maria_M_Portillo-coreas@mcpsmd.org",
    },
    {
        "formal": "Mrs. Wan Li Hsu Chen",
        "title": "Food Svc Satellite Mgr I",
        "email": "WanLi_HsuChen@mcpsmd.org",
    },
]


def xlsx_to_dict(src, sheet=None):
    if sheet is None:
        from openpyxl import load_workbook

        wb = load_workbook(filename=src)
        sheet = wb.active
    else:
        pass

    # col_labels = sheet.row_values(1)
    col_labels = [x.value for x in next(sheet.rows)]

    col_clean_labels = [x.strip() for x in col_labels]
    clean_col = {x: y for x, y in zip(col_labels, col_clean_labels)}

    # num_cols = len(col_labels)

    Directory_Withholding_key = "Directory Withholding-YN"
    if Directory_Withholding_key not in col_labels:
        print(f"{Directory_Withholding_key} was not found. not safe to load this")
        return

    # nuke the preK

    num_withheld = 0
    num_accepted = 0

    # emails_with_includes = {}
    # emails_with_excludes = {}

    for araw in sheet.rows:

        raw_dict = {}
        for label, acell in zip(clean_col, araw):
            raw_dict[label] = str(acell.value)

        adict = {clean_col[x]: str(y) for x, y in raw_dict.items()}
        withheld = False

        delkeys = []
        for k in adict.keys():
            if adict[k] is None:
                delkeys.append(k)
        for k in delkeys:
            del adict[k]
        Directory_Withholding = adict.get(Directory_Withholding_key)

        if Directory_Withholding != "N":
            withheld = True
            # for k in [
            #     "Sch Num",
            #     "School",
            #     "Birth Date",
            #     #'Directory Withholding-YN',
            #     "Phone",
            #     "Address1",
            #     "Address2",
            #     "Relation",
            #     "Name",
            #     "Cell Phone",
            #     "Email",
            #     'Student',
            #         #'Homeroom Teacher', 'Grade',
            # ]:
            #     #adict[k] = withheld_marker
            #     continue
            if Directory_Withholding == "Y":
                pass
            elif "withholding" in Directory_Withholding.lower():
                # withheld = True
                print(f"skipping header based on seeing '{Directory_Withholding}'")
            else:
                print(
                    f"{Directory_Withholding_key} = '{Directory_Withholding}' ... not understood, so dropping this record"
                )

        if not withheld:
            yield adict
            num_accepted += 1
        else:
            num_withheld += 1


@cli.command("make-memberhub-import")
@click.option("--src", help="MCPS export .xlsx", required=True)
@click.pass_context
def make_memberhub_import(ctx, src):
    """setup whatever is needed"""

    extra_org_roles = {
        "victoria.levitas@gmail.com": "Admin, Contact, Member, Customer, Officer",
        "cariaso@gmail.com": "Admin, Officer",
        "chris.press@gmail.com": "Officer",
        # "gillianedick@gmail.com": "Admin, Contact, Member, Customer, Officer",
        # "sarahsandelius@gmail.com": "Admin, Officer",
        # "ckeeling83@gmail.com": "Officer",
        # "jacquelyn_quan@yahoo.com": "Officer",
        # "cherinacyborski@hotmail.com": "Officer",
        # "meghan.holohan@gmail.com": "Officer",
        # "babytrekie@yahoo.com": "Officer",
        # "tajpowell10@gmail.com": "Officer",
        # "wanujarie@gmail.com": "Officer",
        # "katejulian@yahoo.com": "Officer",  # Kate Julian
        # "merritt.m.crowder@mcpsmd.org" :"false",
        # "gabrielle@enrichment-academies.com" :"false",
        # "cynthia_a_houston@mcpsmd.org" :"false",
    }

    errors = False
    fam = {}
    next_fam_id = 123
    seen_fam = {}
    # seen_fam_name = {}
    # seen_fam_name_rev = {}

    grade_teachers = {}

    for adict in xlsx_to_dict(src):
        # print(adict)

        if True:
            fam_id = None
            row = adict

            student_lastname, student_firstname = [
                x.strip() for x in row["Student"].split(",", 2)
            ]
            parent_full = get_relation_name(row)
            parent_lastname, parent_firstname = [
                x.strip() for x in parent_full.split(",", 2)
            ]

            address1, address2 = get_address12(row)
            relation_cell = get_relation_phone(row)
            relation_email = get_relation_email(row)

            fam_val = address1
            if not fam_val.strip():
                fam_val = relation_cell
                if not fam_val.strip():
                    print(f"no family figured out for {fam_val.rstrip()}")
                    breakpoint()

            if fam_val in seen_fam:
                fam_id = seen_fam[fam_val]
            else:
                fam_id = f"fam{next_fam_id}"
                next_fam_id += 1
                seen_fam[fam_val] = fam_id

            # fam_id2fam_name(fam_id, fam_val, name=student_lastname)

            # print(f"##{fam_id}\t{row}##")

            grade = row["Grade"]
            teacher = row["Homeroom Teacher"]
            if grade not in grade_teachers:
                grade_teachers[grade] = {}
            if teacher not in grade_teachers[grade]:
                grade_teachers[grade][teacher] = 0
            grade_teachers[grade][teacher] += 1
            student_info = {
                "grade": row["Grade"],
                "teacher": row["Homeroom Teacher"],
                "firstname": student_firstname,
                "lastname": student_lastname,
                "phone": relation_cell,
                "address1": address1,
                "address2": address2,
            }
            parent_info = {
                "firstname": parent_firstname,
                "lastname": parent_lastname,
                "phone": relation_cell,
                "email": relation_email,  # row.get("Email"),
                "address1": address1,
                "address2": address2,
            }
            parent_errors = False
            if not parent_info["email"]:
                errors = True
                parent_errors = True
                print(f"*** no parent email *** {parent_info}")

            row_info = {
                "student": [
                    student_info,
                ],
                "parent": [],
            }
            if not parent_errors:
                row_info["parent"].append(parent_info)

            if fam_id not in fam:
                fam[fam_id] = row_info
            else:
                if parent_info not in fam[fam_id].get("parent", []):
                    fam[fam_id]["parent"].append(parent_info)
                if student_info not in fam[fam_id]["student"]:
                    fam[fam_id]["student"].append(student_info)

    if True:
        if errors:
            print("please fix the errors above")

        for grade in sorted(grade_teachers):
            for teacher in sorted(grade_teachers[grade]):
                print(f"grade-{grade}-{hub_name(teacher)}")

        if True:
            contacts = []

            for fam_id in fam:

                afam = fam[fam_id]

                afam_name = None

                fam_grades = set()
                fam_teachers = set()

                for astudent in afam["student"]:
                    fam_grades.add(astudent["grade"])

                    if afam_name is None:
                        afam_name = astudent["lastname"]

                    if not astudent["teacher"]:
                        errors = True
                        print(
                            f"*** Warning no teacher for fam_id={fam_id} student={astudent}"
                        )
                    else:
                        fam_teachers.add(astudent["teacher"])

                any_email = False
                for aparent in afam["parent"]:
                    if aparent["email"]:
                        any_email = True
                if not any_email:
                    errors = True
                    print(
                        f"*** Warning no family email for  fam_id={fam_id} student={astudent}"
                    )

                for aparent in afam["parent"]:
                    acontact = {}
                    # RoleName = 'parent_guardian'
                    RoleName = "Contact"

                    if afam_name is None:
                        afam_name = aparent["lastname"]

                    # RoleName = 'Parent/Guardian'
                    ##'RoleName:Year' where:
                    ##* 'RoleName' is the name of the system defined Role ie 'admin', 'parent_guardian', 'contact'
                    ##* 'Year' is an option parameter to define specific year for the user to start & expire.
                    ##* To define multiple roles within the organization, you will add a '+' to separate the difference roles ie
                    ##'Admin:2022+Parent/Guardian'
                    ##!!!## this example is in capitalization (admin vs Admin ; Parent/Guardian vs parent_guardian)

                    ##### Hubs & Hub Roles
                    ##There are many roles that a User can be assigned in Memberhub, both roles for a User in an
                    ##Organization, in a Hub, and in a Family. The general format for a Role is +HubName:RoleName:Year.
                    ##* '+' is required when assigning more than one hub role to a user
                    ##* 'HubName' is the name of the Hub
                    ##* 'RoleName' is the name of the role
                    ##* 'Year' is optional and may be included to set the date that the role is to expire. If :Year is not provided,
                    ##there will be no expiration of the role. Expiry can be expressed as :yyyy (a 4 digit year which will be
                    ##interpreted as yyyy-06-30).
                    ##An example of adding to roles to a single hub, plus an additional role to a 2nd hub:
                    ##'Hub1:Admin:2022+Hub1:Parent/Guadian+Hub2:Parent/Guardian'
                    hubs = set()
                    for grade in fam_grades:
                        ahub_name = hub_name(f"grade-{grade}")
                        ahub_role = RoleName
                        ahub_year = str(END_YEAR)
                        ahub = ":".join([ahub_name, ahub_role, ahub_year])
                        hubs.add(ahub)
                    for teacher in fam_teachers:
                        if teacher:
                            ahub_name = hub_name(f"teacher-{teacher}")
                            ahub_role = RoleName
                            ahub_year = str(END_YEAR)
                            ahub = ":".join([ahub_name, ahub_role, ahub_year])
                            hubs.add(ahub)
                    hub_str = "+".join(hubs)
                    acontact["Hubs"] = hub_str
                    # acontact['Organization Role'] = hub_str

                    acontact["First Name"] = aparent["firstname"]
                    acontact["Last Name"] = aparent["lastname"]
                    acontact["Email"] = aparent["email"]
                    acontact["Phone Number"] = aparent["phone"]

                    acontact["Address"] = aparent["address1"]

                    if m := re.search(r"([\w\s+]+), (\w\w) (\d+)", aparent["address2"]):

                        city, state, zipcode = m.groups()
                        city = city.strip()
                    acontact["City"] = city
                    acontact["State"] = state
                    acontact["Zip"] = zipcode

                    acontact["Family Name"] = normalize(afam_name, fam_id)
                    acontact["Family Role"] = "Parent/Guardian"
                    # acontact["Family Role"] = "Contact"

                    if aparent.get("email") in extra_org_roles:
                        acontact["Organization Role"] = "Admin"
                    else:
                        acontact["Organization Role"] = "Contact"

                    # if not acontact['Email'])
                    #    print(f"*** WILL NOT LOAD contact with no Email {acontact}")

                    contacts.append(acontact)

                for astudent in afam["student"]:
                    hubs = set()
                    acontact = {}

                    RoleName = "Student"

                    grade = astudent["grade"]
                    ahub_name = hub_name(f"grade-{grade}")
                    ahub_role = RoleName
                    ahub_year = str(END_YEAR)
                    ahub = ":".join([ahub_name, ahub_role, ahub_year])
                    hubs.add(ahub)

                    teacher = astudent["teacher"]
                    if teacher:
                        ahub_name = hub_name(f"teacher-{teacher}")
                        ahub_role = RoleName
                        ahub_year = str(END_YEAR)
                        ahub = ":".join([ahub_name, ahub_role, ahub_year])
                        hubs.add(ahub)
                    else:
                        print(f"*** Warning no teacher for {astudent}")

                    hub_str = "+".join(hubs)
                    acontact["Hubs"] = hub_str
                    # acontact['Organization Role'] = hub_str

                    acontact["First Name"] = astudent["firstname"]
                    acontact["Last Name"] = astudent["lastname"]
                    # acontact['Email'] = astudent['email']
                    # acontact['Phone Number'] = astudent['phone']

                    acontact["Address"] = astudent["address1"]

                    if m := re.search(
                        r"([\w\s+]+), (\w\w) (\d+)", astudent["address2"]
                    ):

                        city, state, zipcode = m.groups()
                        city = city.strip()
                    acontact["City"] = city
                    acontact["State"] = state
                    acontact["Zip"] = zipcode
                    # acontact["Family Name"] = str(fam_id)
                    acontact["Family Name"] = normalize(afam_name, fam_id)

                    acontact["Organization Role"] = "Student"
                    acontact["Family Role"] = "Child"

                    contacts.append(acontact)

        if True:
            contact_keys = [
                "First Name",
                "Last Name",
                "Email",
                "Phone Number",
                "Organization Role",
                "Address",
                "City",
                "State",
                "Zip",
                "Family Name",
                "Family Role",
                "Hubs",
                "Contact Property 1",
                "Contact Property 2",
                # "Family Name",
                # "Family Role",
                # "Address",
                # "First Name",
                # "Last Name",
                # "City",
                # "State",
                # "Zip",
                # "Email",
                #
                #'Phone Number','Organization Role','Hubs',
            ]
            for contact in contacts:
                for k in contact:
                    if k not in contact_keys:
                        contact_keys.append(k)

            seen_emails = set()
            for contact in contacts:
                if contact.get("Organization Role") not in [
                    "Contact",
                    "Admin",
                    "Store Admin",
                    "Student",
                ]:
                    print(f"bad Organization Roles for {contact}")

                if contact.get("Organization Role") not in ["Student"]:
                    if contact.get("Email") in seen_emails:
                        print(f"bad duplicate email {contact.get('Email')}")
                    else:
                        seen_emails.add(contact.get("Email"))

            with open("ready_to_load.csv", "w") as outfh:
                outfh.write(",".join(contact_keys))
                outfh.write("\n")
                for contact in contacts:
                    out = []
                    for key in contact_keys:
                        val = ""
                        if key in contact:
                            val = contact[key]
                            if val:
                                if "," in val:
                                    print("*** Warning comma in planned output {val}")
                                    breakpoint()
                                if '"' in val:
                                    print(
                                        "*** Warning double-quote in planned output {val}"
                                    )
                                    breakpoint()
                            if val is None:
                                val = ""
                        out.append(val)
                    outfh.write(",".join(out))
                    outfh.write("\n")

            print("*** YOU MUST manually make these hubs ***")
            for ahub_name in sorted(list(all_hubs)):
                print(ahub_name)

            # ALL CONTACTS MUST HAVE AN EMAIL WITH THEIR PROFILE IN ORDER TO BE IMPORTED WITH
            # THE EXCEPTION OF ANY CONTACT WITH THE ROLE STUDENT OR CHILD.

            ### Organization Role

            ##Users can be assigned multiple roles for the Organization itself. The general structure for these roles are


def get_relation_name(entry):
    relation_name = entry.get("Parent/Guardian Name") or entry.get("Name")
    return relation_name


def get_address12(entry):

    address1 = None
    address2 = None
    if entry.get("Home Address1"):
        address1 = entry.get("Home Address1")
        address2 = entry.get("Home Address2")
    elif entry.get("Mailing Address1"):
        address1 = entry.get("Mailing Address1")
        address2 = entry.get("Mailing Address2")
    elif entry.get("Address1"):
        address1 = entry.get("Address1")
        address2 = entry.get("Address2")
    return address1, address2


def get_relation_phone(entry):
    relation_cell = (
        entry.get("Parent/Guardian Cell Phone")
        or entry.get("Cell Phone")
        or entry.get("Phone")
    )
    return relation_cell


def get_relation_email(entry):
    val = entry.get("Parent/Guardian Email") or entry.get("Email")
    return val


END_YEAR = 2025


all_hubs = set()


def hub_name(s):
    if s == "teacher-":
        breakpoint()
        return None
    if "+" in s:
        breakpoint()
    s = s.replace(", ", "-")
    s = s.replace(",", "-")
    s = s.replace(" ", "-")
    all_hubs.add(s)
    return s


seen_long_fams = {}
seen_short_fams = {}
seen_complex = {}


def normalize(name, fam_id):
    if name not in seen_complex:
        seen_complex[name] = {fam_id: 1}

    if fam_id not in seen_complex[name]:
        idx = max(seen_complex[name].values()) + 1
        seen_complex[name][fam_id] = idx

    idx = seen_complex[name][fam_id]
    if idx == 1:
        long_name = name
    else:
        long_name = f"{name}#{idx}"
    return long_name


def as_email(username, recipients, attachment):

    sender_email = username

    message = MIMEMultipart()
    message["Subject"] = "SomersetPTA Directory Fall 2024"
    message["From"] = f"SomersetPTA Directory <{sender_email}>"
    message["To"] = ", ".join(recipients)
    # message.preamble = "preamble"

    text_body = f"""This Somerset ES PTA Directory has been made just for you {recipients[0]}. We hope you'll find it useful.
    """
    body = MIMEText(text_body)
    message.attach(body)

    if attachment:
        with open(attachment, "rb") as attachment_fh:
            part = MIMEBase("application", "pdf")
            part.set_payload(attachment_fh.read())
        encoders.encode_base64(part)
        attachment_name = pathlib.Path(attachment).name
        part.add_header(
            "Content-Disposition",
            f'attachment; filename="{attachment_name}"',
        )
        message.attach(part)

    # print(f"{len(message.as_string()):,d} bytes")

    yield message


def send_emails(username, password, messages):

    sender_email = username
    sender_password = password

    with smtplib.SMTP_SSL("smtp.gmail.com", 465) as server:
        a1 = server.login(sender_email, sender_password)
        print(a1)
        # a2 = server.sendmail(sender_email, recipient, message.as_string())
        # 1) You should use msg.as_string() if you call smtplib.SMTP.sendmail(). Alternatively, if you have Python 3.2 or newer, you can use
        for message in messages:
            # print(message.as_string())
            a2 = server.send_message(message)
            print(a2)


@cli.command("send-test-directory-email")
@click.pass_context
def send_test_directory_email(ctx):
    """setup whatever is needed"""

    login_username = "directory@somersetpta.org"
    from dotenv import load_dotenv

    load_dotenv()
    password = os.environ["SOMERSETPTA_DIRECTORY_PASSWORD"]
    recipients = ["cariaso@gmail.com"]
    stream = as_email(
        username=login_username,
        recipients=recipients,
        # attachment=filename,
    )
    send_emails(username=login_username, password=password, messages=stream)


if __name__ == "__main__":
    cli()
